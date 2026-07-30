import eventlet

eventlet.monkey_patch()

import hmac
import logging
import logging.handlers
from datetime import UTC, datetime, timedelta
from functools import wraps

from flask import (
    Flask,
    Response,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_cors import CORS
from flask_socketio import SocketIO, emit

from database import db_session, init_db
from models import (
    Approval,
    AuditLog,
    Device,
    Employee,
    Equipment,
    GateLog,
    GateMapping,
    SiteSetting,
    Vehicle,
    Visitor,
)

# Import shared utilities and extensions to avoid circular imports
from utils import _utcnow, login_required, role_required, require_api_key, log_audit
from extensions import (
    __version__,
    ENABLE_AI_CHAT,
    limiter,
    socketio,
    logger,
    OLLAMA_BASE_URL,
    OLLAMA_MODEL,
    OLLAMA_MODEL_FULL,
    _ollama_provider,
    _ollama_available,
    _ollama_checked,
    _check_ollama,
    PSUTIL_AVAILABLE,
    metrics_history,
    request_timestamps,
    parse_log_line,
)


# _utcnow is imported from utils


import hashlib
import io
import json
import os
import re
import select
import socket
import subprocess
import sys
import threading
import time

import openpyxl
import qrcode
import requests
from barcode.codex import Code128
from barcode.writer import ImageWriter

# PSUTIL_AVAILABLE is imported from extensions
import base64
import re as _re

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import inch, landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func

# Sanitize strings for openpyxl (strip illegal XML characters)
_ILLEGAL_XML_CHARS = _re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ud800-\udfff\ufdd0-\ufdef\ufffe\uffff]"
)


def _sanitize_cell(value):
    """Remove illegal XML characters that openpyxl cannot write."""
    if isinstance(value, str):
        return _ILLEGAL_XML_CHARS.sub("", value)
    return value


# Load environment variables from .env file
try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, use system environment variables

# ------------------- Structured Logging -------------------
_LOG_LEVEL = os.environ.get("LOG_LEVEL", "INFO").upper()
_LOG_FILE = os.environ.get("LOG_FILE", "")

logging.basicConfig(
    level=getattr(logging, _LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
# logger is imported from extensions

if _LOG_FILE:
    _handler = logging.handlers.RotatingFileHandler(
        _LOG_FILE, maxBytes=10 * 1024 * 1024, backupCount=5
    )
    _handler.setFormatter(
        logging.Formatter(
            "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    logger.addHandler(_handler)

# ------------------- App Init -------------------
_secret_key = os.environ.get("SECRET_KEY")
_is_production = os.environ.get("FLASK_ENV", "production").lower() == "production"
if not _secret_key:
    if _is_production:
        raise RuntimeError(
            "SECRET_KEY environment variable is not set. "
            'Generate one with: python -c "import secrets; print(secrets.token_hex(32))" '
            "and add it to your .env file."
        )
    _secret_key = os.urandom(24).hex()
    logger.warning(
        "SECRET_KEY not set — using random key (sessions reset on restart). Set SECRET_KEY in .env"
    )

# Warn at startup if HARDWARE_API_KEY is not configured
_hardware_key_check = os.environ.get("HARDWARE_API_KEY")
if not _hardware_key_check:
    logger.warning(
        "HARDWARE_API_KEY is not set — hardware API key authentication is disabled. "
        "Set HARDWARE_API_KEY in .env for production deployments."
    )

app = Flask(__name__)
app.secret_key = _secret_key
app.permanent_session_lifetime = timedelta(minutes=30)
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("HTTPS", "false").lower() == "true"
app.config["WTF_CSRF_TIME_LIMIT"] = 3600  # 1 hour CSRF token validity
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 43200  # 12 hour static asset cache

# Response compression (gzip/deflate)
from flask_compress import Compress

Compress(app)

# CSRF Protection - Manual implementation
# API routes (/api/*) are exempt because they use header-based authentication (X-API-Key)
# instead of session-based CSRF tokens. Web routes use session cookies and require CSRF.
from flask_wtf.csrf import CSRFError, CSRFProtect

app.config["WTF_CSRF_CHECK_DEFAULT"] = False  # Manual CSRF checking
csrf = CSRFProtect(app)


@app.before_request
def csrf_protect_non_api():
    """Apply CSRF check to non-API state-changing requests.

    API routes (/api/*) use X-API-Key header authentication and are exempt from CSRF.
    Web routes use session cookies and require CSRF tokens for POST/PUT/PATCH/DELETE.
    """
    if app.config.get("TESTING"):
        return
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        # Exempt /api/ routes - they use header-based auth, not session cookies
        if not request.path.startswith("/api/"):
            csrf.protect()


@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    return render_template(
        "login.html", error="Session expired. Please try again."
    ), 400


@app.after_request
def add_security_headers(response):
    """Add security headers to all responses."""
    # Content-Security-Policy (allow inline for CDN-loaded scripts)
    csp = "default-src 'self'; script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://cdn.socket.io https://cdnjs.cloudflare.com https://unpkg.com; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; font-src 'self' https://fonts.gstatic.com; img-src 'self' data:; connect-src 'self' https://cdn.socket.io"
    response.headers["Content-Security-Policy"] = csp
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Strict-Transport-Security"] = (
        "max-age=31536000; includeSubDomains"
    )
    return response


# Rate Limiting
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# SECURITY NOTE: storage_uri="memory://" means rate limit counters are stored in-process
# and reset on every app restart. In production, this means:
#   1. An attacker can bypass rate limits by triggering app restarts (if they have access)
#   2. Multi-worker deployments won't share rate limit state
# For production with multiple workers, use Redis: storage_uri="redis://localhost:6379"
_ratelimit_storage = os.environ.get("REDIS_URL") or os.environ.get(
    "RATELIMIT_STORAGE_URI", "memory://"
)
# Initialize limiter with app and custom storage
limiter.init_app(app)
limiter.storage_uri = _ratelimit_storage
if _ratelimit_storage == "memory://":
    logger.warning(
        "Rate limiting uses in-memory storage (resets on restart). "
        "For production with multiple workers, configure REDIS_URL or RATELIMIT_STORAGE_URI."
    )

# Enable CORS for API endpoints (required for mobile scanner access)
# CORS configuration - restricted to production origins in production
# For development/debug allow localhost, production should be explicit
_cors_origins = os.environ.get(
    "CORS_ORIGINS", "*"
)  # Set CORS_ORIGINS env var to restrict
CORS(
    app,
    resources={
        r"/api/*": {
            "origins": _cors_origins.split(",") if _cors_origins != "*" else "*",
            "methods": ["GET", "POST", "OPTIONS"],
            "allow_headers": ["Content-Type", "X-API-Key", "X-CSRFToken"],
            "supports_credentials": True,
        }
    },
)

# Initialize SocketIO with app
socketio.init_app(app, cors_allowed_origins=_cors_origins.split(",") if _cors_origins != "*" else "*")


# Custom Jinja2 filters
@app.template_filter("from_json")
def from_json_filter(value):
    """Parse JSON string to Python object."""
    if not value:
        return {}
    try:
        return json.loads(value)
    except (json.JSONDecodeError, ValueError):
        return {}


# Initialize database
init_db()


# ------------------- JSON Error Handlers -------------------
@app.errorhandler(404)
def json_404(error):
    return jsonify(
        {"error": "Not found", "message": "The requested resource was not found"}
    ), 404


@app.errorhandler(405)
def json_405(error):
    return jsonify(
        {
            "error": "Method not allowed",
            "message": "This HTTP method is not allowed for this endpoint",
        }
    ), 405


@app.errorhandler(500)
def json_500(error):
    return jsonify({"error": "Internal server error", "message": str(error)}), 500


# ------------------- Ollama Local AI Configuration (100% Free Endpoint) -------------------
# ENABLE_AI_CHAT is imported from extensions
app.config["ENABLE_AI_CHAT"] = ENABLE_AI_CHAT

# Initialize Ollama configuration from environment variables
from extensions import init_ollama_config
init_ollama_config(
    base_url=os.environ.get("OLLAMA_URL", "http://localhost:11434"),
    model=os.environ.get("OLLAMA_MODEL", "mine-assistant-fast"),
    model_full=os.environ.get("OLLAMA_MODEL_FULL", "mine-assistant"),
    provider="local",
    available=False
)

# In-memory metrics storage and request tracking are imported from extensions

# _check_ollama is imported from extensions

# Try at startup (non-blocking if Ollama isn't ready yet)
_check_ollama()


# ------------------- Multi-Port Scanner Listener -------------------
# Listens on multiple UDP/TCP ports to catch scanners with various configurations
from services.listeners import (
    _ensure_device_exists,
)
from services.listeners import (
    init_all_scanner_listeners as _init_listeners,
)
from services.listeners import (
    process_scan_data as _process_scan_data_listener,
)


def process_scan_data(qr_data, source_ip, protocol="UDP"):
    return _process_scan_data_listener(
        qr_data, source_ip, protocol=protocol, process_qr_callback=_process_qr_scan, socketio_instance=socketio
    )

def init_all_scanner_listeners():
    _init_listeners(process_qr_callback=_process_qr_scan, socketio_instance=socketio)


# ------------------- Decorators -------------------
# login_required, role_required, and require_api_key are imported from utils


# ------------------- Public routes -------------------
@app.route("/")
def index():
    if not session.get("logged_in"):
        return redirect(url_for("auth.login"))
    return redirect(url_for("dashboard.dashboard"))


# ------------------- Visitor QR Request (Public) -------------------
@app.route("/visitor_request", methods=["GET", "POST"])
def visitor_request():
    """Public page: visitors can request a QR pass without logging in."""
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        company = request.form.get("company", "").strip()
        purpose = request.form.get("purpose", "").strip()
        meeting_person = request.form.get("meeting_person", "").strip()

        if not name:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"success": False, "error": "Visitor name is required."})
            return render_template(
                "visitor_request.html", error="Visitor name is required."
            )

        # Validate department PIN
        submitted_pin = request.form.get("pin", "").strip()
        pin_setting = (
            db_session.query(SiteSetting).filter_by(key="visitor_request_pin").first()
        )
        expected_pin = pin_setting.value if pin_setting else "1234"
        if submitted_pin != expected_pin:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"success": False, "error": "Invalid PIN. Contact your HOD or admin."})
            return render_template(
                "visitor_request.html", error="Invalid PIN. Contact your HOD or admin."
            )

        visitor = Visitor(
            name=name,
            company=company,
            purpose=purpose,
            meeting_person=meeting_person,
            status="Pending Approval",
        )
        db_session.add(visitor)
        db_session.flush()

        # Generate QR code hash (same pattern as generate_qr_code)
        qr_data = f"VIS:{visitor.id}:{visitor.name}:{datetime.now().timestamp()}"
        qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:32].upper()
        visitor.qr_code = qr_hash

        # Create Approval record so admins can approve
        approval = Approval(
            request_type="Visitor QR Request",
            request_id=visitor.id,
            requester_name=name,
            details=f"Visitor: {name} | Company: {company} | Reason: {purpose} | Meeting: {meeting_person}",
            status="Pending",
            target_table="visitors",
        )
        db_session.add(approval)
        db_session.commit()

        socketio.emit("visitor_checkin", {"name": visitor.name})
        socketio.emit(
            "stats_update",
            {
                "pending_approvals": db_session.query(Approval)
                .filter_by(status="Pending")
                .count()
            },
        )

        # Generate QR image as base64 for display
        qr = qrcode.QRCode(version=4, box_size=10, border=4)
        qr.add_data(qr_hash)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        qr_img.save(buf, format="PNG")
        qr_base64 = base64.b64encode(buf.getvalue()).decode("utf-8")

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({
                "success": True,
                "qr_base64": qr_base64,
                "visitor": {
                    "name": visitor.name,
                    "company": visitor.company,
                    "purpose": visitor.purpose,
                    "meeting_person": visitor.meeting_person,
                },
            })

        return render_template(
            "visitor_request.html",
            success=True,
            visitor=visitor,
            qr_base64=qr_base64,
        )

    return render_template("visitor_request.html")


# ------------------- Emergency Muster -------------------
@app.route("/muster")
@login_required
def emergency_muster():
    """Emergency muster roll: shows everyone currently on-site based on gate logs."""
    from sqlalchemy import and_, func

    # Subquery: latest gate log per entity
    latest_log_subq = (
        db_session.query(
            GateLog.access_type,
            GateLog.entity_id,
            func.max(GateLog.scanned_at).label("last_scan"),
        )
        .filter(GateLog.access_granted)
        .group_by(GateLog.access_type, GateLog.entity_id)
        .subquery()
    )

    # Get the actual log entries for those latest scans
    latest_logs = (
        db_session.query(GateLog)
        .join(
            latest_log_subq,
            and_(
                GateLog.access_type == latest_log_subq.c.access_type,
                GateLog.entity_id == latest_log_subq.c.entity_id,
                GateLog.scanned_at == latest_log_subq.c.last_scan,
            ),
        )
        .filter(GateLog.direction == "IN")
        .all()
    )

    on_site = {
        "employees": [],
        "visitors": [],
        "vehicles": [],
    }

    # Pre-fetch all employee data in one query (fixes N+1)
    emp_ids = [
        log.entity_id
        for log in latest_logs
        if log.access_type == "employee" and log.entity_id
    ]
    emp_map = {}
    if emp_ids:
        emps = db_session.query(Employee).filter(Employee.id.in_(emp_ids)).all()
        emp_map = {e.id: e for e in emps}

    for log in latest_logs:
        entry = {
            "name": log.entity_name,
            "id": log.entity_id,
            "gate": log.gate_location or "Unknown",
            "time_in": log.scanned_at,
        }
        if log.access_type == "employee":
            emp = emp_map.get(log.entity_id)
            if emp:
                entry["job_title"] = emp.job_title or "N/A"
                entry["emp_code"] = emp.emp_code
            on_site["employees"].append(entry)
        elif log.access_type == "visitor":
            on_site["visitors"].append(entry)
        elif log.access_type == "vehicle":
            on_site["vehicles"].append(entry)

    total_on_site = (
        len(on_site["employees"]) + len(on_site["visitors"]) + len(on_site["vehicles"])
    )

    return render_template(
        "muster.html",
        on_site=on_site,
        total_on_site=total_on_site,
    )


# ------------------- Device Onboarding Terminal -------------------
@app.route("/onboard")
def onboard():
    """Device onboarding terminal with QR code for provisioning"""
    # Get server IP
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        server_ip = s.getsockname()[0]
        s.close()
    except Exception:
        server_ip = "127.0.0.1"

    # Get request values
    server_port = request.args.get("port", "8080")

    # Build Config JSON (standard format for mobile auto-config)
    config_payload = {
        "server_ip": server_ip,
        "server_port": server_port,
        "api_endpoint": f"http://{server_ip}:{server_port}/api/scanner/receive",
        "api_key": "MINE-CONFIG-ABC-123",
        "timestamp": datetime.now().isoformat(),
    }
    config_json = json.dumps(config_payload)

    # Generate Config QR code
    qr_config = qrcode.QRCode(version=4, box_size=10, border=4)
    qr_config.add_data(config_json)
    qr_config.make(fit=True)
    img_config = qr_config.make_image(fill_color="black", back_color="white")

    img_buffer_config = io.BytesIO()
    img_config.save(img_buffer_config, format="PNG")
    img_buffer_config.seek(0)
    config_qr_image = f"data:image/png;base64,{base64.b64encode(img_buffer_config.getvalue()).decode()}"

    # Generate App Download QR code
    app_download_url = f"http://{server_ip}:{server_port}" + url_for(
        "static", filename="downloads/QrMobile.apk"
    )
    config_url = f"http://{server_ip}:{server_port}/api/config/infowedge"
    qr_app = qrcode.QRCode(version=4, box_size=10, border=4)
    qr_app.add_data(app_download_url)
    qr_app.make(fit=True)
    img_app = qr_app.make_image(fill_color="black", back_color="white")

    img_buffer_app = io.BytesIO()
    img_app.save(img_buffer_app, format="PNG")
    img_buffer_app.seek(0)
    app_qr_image = (
        f"data:image/png;base64,{base64.b64encode(img_buffer_app.getvalue()).decode()}"
    )

    # Get device stats
    total_devices = db_session.query(Device).count()
    active_devices = db_session.query(Device).filter(Device.status == "online").count()
    total_scans = (
        db_session.query(Device).with_entities(func.sum(Device.total_scans)).scalar()
        or 0
    )

    # Get recent devices
    recent_devices = (
        db_session.query(Device).order_by(Device.last_seen.desc()).limit(5).all()
    )
    recent = []
    for d in recent_devices:
        recent.append(
            {
                "device_name": d.device_name,
                "ip_address": d.ip_address or d.mac_address or "Unknown",
                "last_seen": d.last_seen.strftime("%H:%M:%S")
                if d.last_seen
                else "Never",
            }
        )

    stats = {
        "total_devices": total_devices,
        "active_devices": active_devices,
        "total_scans": total_scans,
    }

    return render_template(
        "onboard.html",
        config_qr_image=config_qr_image,
        app_qr_image=app_qr_image,
        app_download_url=app_download_url,
        config_url=config_url,
        server_ip=server_ip,
        server_port=server_port,
        stats=stats,
        recent_devices=recent,
    )


def _extract_scan_fields(payload: dict) -> dict:
    """Return a dict with only keys that look like scan-related data."""
    # Common keys used by various scanner apps and wedges
    keywords = [
        "scan",
        "code",
        "qr",
        "barcode",
        "data",
        "text",
        "value",
        "result",
        "raw",
        "msg",
        "content",
        "info",
    ]
    scan_keys = [k for k in payload if any(sub in k.lower() for sub in keywords)]
    return {k: payload[k] for k in scan_keys}


@app.route("/api/config/infowedge")
def get_infowedge_config():
    """Return InfoWedge configuration for download"""
    # Auto-create device for the requesting IP
    client_ip = request.remote_addr
    _ensure_device_exists(client_ip)

    # Get server IP
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        server_ip = s.getsockname()[0]
        s.close()
    except Exception:
        server_ip = "127.0.0.1"

    server_port = request.args.get("port", "8080")

    # Create config content (simple text file with instructions)
    config_content = f"""# InfoWedge Configuration
# Auto-generated for Arch-System

[Network]
Server IP: {server_ip}
Server Port: {server_port}
Protocol: UDP

[Basic Settings]
Enable Basic Data Formatting: true
Send ENTER Key: true
Enable Scanner: true
"""

    return Response(
        config_content,
        mimetype="text/plain",
        headers={"Content-Disposition": "attachment; filename=infowedge_config.txt"},
    )


@app.route("/api/device/register", methods=["POST"])
def register_device():
    """Register a new device when it connects"""
    data = request.get_json() or {}

    device_name = data.get(
        "device_name", f"Device-{data.get('mac_address', 'unknown')[:8]}"
    )
    mac_address = data.get("mac_address")
    ip_address = data.get("ip_address") or request.remote_addr
    device_type = data.get("device_type", "Unknown")

    # Check if device already exists
    existing = None
    if mac_address:
        existing = db_session.query(Device).filter_by(mac_address=mac_address).first()
    elif ip_address:
        existing = db_session.query(Device).filter_by(ip_address=ip_address).first()

    if existing:
        # Update existing device
        existing.last_seen = _utcnow()
        existing.status = "online"
        if ip_address:
            existing.ip_address = ip_address
        msg = "Device updated"
    else:
        # Create new device
        device = Device(
            device_name=device_name,
            device_type=device_type,
            mac_address=mac_address,
            ip_address=ip_address,
            status="online",
        )
        db_session.add(device)
        db_session.commit()
        msg = "Device registered"

    db_session.commit()

    return jsonify({"success": True, "message": msg})


@app.route("/api/device/heartbeat", methods=["POST"])
def device_heartbeat():
    """Update device last_seen and scan count"""
    data = request.get_json() or {}

    ip_address = data.get("ip_address") or request.remote_addr
    scans = data.get("scans", 0)

    device = db_session.query(Device).filter_by(ip_address=ip_address).first()
    if device:
        device.last_seen = _utcnow()
        device.status = "online"
        device.total_scans += scans
        db_session.commit()

    return jsonify({"success": True})


@app.route("/api/device/confirm", methods=["POST"])
def confirm_device():
    """Confirm a pending device"""
    data = request.get_json() or {}
    device_id = data.get("device_id")
    device_name = data.get("device_name", "Chainway C66")
    device_type = data.get("device_type", "C66")

    device = db_session.query(Device).filter_by(id=device_id).first()
    if device:
        device.device_name = device_name
        device.device_type = device_type
        device.status = "online"
        db_session.commit()
        return jsonify({"success": True, "message": "Device confirmed"})
    return jsonify({"success": False, "message": "Device not found"})


@app.route("/api/device/reject", methods=["POST"])
def reject_device():
    """Reject and remove a pending device"""
    data = request.get_json() or {}
    device_id = data.get("device_id")

    device = db_session.query(Device).filter_by(id=device_id).first()
    if device:
        db_session.delete(device)
        db_session.commit()
        return jsonify({"success": True, "message": "Device removed"})
    return jsonify({"success": False, "message": "Device not found"})


# Remove duplicate dashboard route definition below - keeping only this one


# ------------------- WebSocket -------------------
_stats_cache = {"data": None, "ts": 0}


@socketio.on("request_stats")
def handle_stats_request():
    import time

    now = time.time()
    if _stats_cache["data"] and (now - _stats_cache["ts"]) < 5:
        emit("stats_update", _stats_cache["data"])
        return
    stats = {
        "employees": db_session.query(Employee).count(),
        "vehicles": db_session.query(Vehicle).count(),
        "visitors": db_session.query(Visitor).filter_by(status="Checked In").count(),
        "pending_approvals": db_session.query(Approval)
        .filter_by(status="Pending")
        .count(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _stats_cache["data"] = stats
    _stats_cache["ts"] = now
    emit("stats_update", stats)


# ------------------- Approvals -------------------
@app.route("/pending_approvals")
@login_required
def pending_approvals():
    approvals = db_session.query(Approval).filter_by(status="Pending").all()

    # Decode scanned QR data for each approval
    decoded_map = {}
    for appr in approvals:
        if appr.scanned_data:
            try:
                stored = json.loads(appr.scanned_data)
                raw_qr = (
                    stored.get("qr_code")
                    or stored.get("raw_data")
                    or stored.get("original_data")
                    or ""
                )
                decoded = decode_qr_data(raw_qr)
                # Merge stored fields over decoded (stored takes precedence)
                for k in ("employee_id", "name", "position", "department", "area"):
                    if stored.get(k) and not decoded.get(k):
                        decoded[k] = stored[k]
                decoded_map[appr.id] = decoded
            except (json.JSONDecodeError, ValueError):
                decoded_map[appr.id] = decode_qr_data(appr.scanned_data)
        else:
            decoded_map[appr.id] = {"raw_data": None, "format": "none"}

    return render_template(
        "pending_approvals.html", approvals=approvals, decoded_map=decoded_map
    )


@app.route("/api/approval/<int:id>")
@login_required
def get_approval(id):
    approval = db_session.query(Approval).filter_by(id=id).first()
    if not approval:
        return jsonify({"error": "Not found"}), 404

    # Build response with basic info
    response = {
        "id": approval.id,
        "request_type": approval.request_type,
        "request_id": approval.request_id,
        "requester_name": approval.requester_name,
        "details": approval.details,
        "created_at": approval.created_at.strftime("%Y-%m-%d %H:%M"),
        "status": approval.status,
        "target_table": approval.target_table,
        "scanned_data": json.loads(approval.scanned_data)
        if approval.scanned_data
        else None,
    }

    # Add entity-specific details
    if approval.request_type == "Employee" and approval.request_id:
        employee = db_session.query(Employee).filter_by(id=approval.request_id).first()
        if employee:
            response["entity_data"] = {
                "id": employee.id,
                "emp_code": employee.emp_code,
                "initials": employee.initials or "N/A",
                "first_name": employee.first_name,
                "second_name": employee.second_name or "N/A",
                "surname": employee.surname,
                "id_number": employee.id_number,
                "job_title": employee.job_title or "N/A",
                "induction": employee.induction or "N/A",
                "induction_expiry": employee.induction_expiry.strftime("%Y-%m-%d")
                if employee.induction_expiry
                else "N/A",
                "medical": employee.medical or "N/A",
                "medical_expiry": employee.medical_expiry.strftime("%Y-%m-%d")
                if employee.medical_expiry
                else "N/A",
                "status": employee.status,
            }

    elif approval.request_type == "Vehicle" and approval.request_id:
        vehicle = db_session.query(Vehicle).filter_by(id=approval.request_id).first()
        if vehicle:
            response["entity_data"] = {
                "id": vehicle.id,
                "fleet_id": vehicle.fleet_id,
                "registration_expiry": vehicle.registration_expiry.strftime("%Y-%m-%d")
                if vehicle.registration_expiry
                else "N/A",
                "status": vehicle.status,
            }

    elif approval.request_type == "Visitor" and approval.request_id:
        visitor = db_session.query(Visitor).filter_by(id=approval.request_id).first()
        if visitor:
            host_name = visitor.host.name if visitor.host else "N/A"
            response["entity_data"] = {
                "id": visitor.id,
                "name": visitor.name,
                "company": visitor.company or "N/A",
                "purpose": visitor.purpose or "N/A",
                "host_name": host_name,
                "check_in_time": visitor.check_in_time.strftime("%Y-%m-%d %H:%M")
                if visitor.check_in_time
                else "N/A",
                "status": visitor.status,
            }

    return jsonify(response)


@app.route("/approve_request/<int:id>", methods=["POST"])
@login_required
@role_required(["admin", "manager"])
def approve_request(id):
    approval = db_session.query(Approval).filter_by(id=id).first()
    if approval:
        data = request.get_json() or {}
        approval.status = "Approved"
        approval.approved_by = session.get("username")
        approval.approval_date = _utcnow()
        approval.comments = data.get("comment", "")
        approval.target_table = data.get(
            "target_table", "employees"
        )  # 'employees' or 'fleet'

        # Get scanned data if available
        scanned_data = {}
        if approval.scanned_data:
            try:
                scanned_data = json.loads(approval.scanned_data)
            except Exception:
                scanned_data = {}

        # Also get form data for new records
        form_data = data.get("form_data", {})

        new_entity = None
        entity_type = None

        # Create new record based on target_table
        if approval.target_table == "employees":
            original_qr_code = scanned_data.get("qr_code") or scanned_data.get(
                "original_data"
            )
            emp_id_from_scan = scanned_data.get("employee_id")

            # If no employee_id in scanned_data, try to extract it from qr_code
            if not emp_id_from_scan and original_qr_code:
                import re

                id_match = re.search(r"ID[:\s]*(\d+)", original_qr_code)
                if id_match:
                    emp_id_from_scan = id_match.group(1)

            # Generate temp ID if still not found
            if not emp_id_from_scan:
                emp_id_from_scan = f"TEMP{approval.id:08d}"

            existing = None
            if original_qr_code:
                existing = (
                    db_session.query(Employee)
                    .filter_by(qr_code=original_qr_code)
                    .first()
                )
            if not existing and emp_id_from_scan:
                existing = (
                    db_session.query(Employee)
                    .filter_by(emp_code=emp_id_from_scan)
                    .first()
                )

            if not existing:
                # Create new employee - use original QR code from scan so next scan matches
                # Parse name into first_name and surname
                full_name = (
                    scanned_data.get("name")
                    or form_data.get("name")
                    or approval.requester_name
                    or "Unknown"
                )
                name_parts = full_name.split(None, 1)
                first_name = name_parts[0] if name_parts else "Unknown"
                surname = name_parts[1] if len(name_parts) > 1 else ""

                new_employee = Employee(
                    emp_code=emp_id_from_scan,
                    first_name=first_name,
                    surname=surname,
                    id_number=scanned_data.get("id_number")
                    or form_data.get("id_number")
                    or emp_id_from_scan,
                    job_title=scanned_data.get("position")
                    or form_data.get("position")
                    or "Unknown",
                    status="Active",
                    qr_code=original_qr_code,  # Use original QR so next scan matches
                )
                db_session.add(new_employee)
                db_session.flush()  # Get the ID

                new_entity = new_employee
                entity_type = "employee"
                approval.request_id = new_employee.id
            else:
                # Update existing employee to Active
                existing.status = "Active"
                new_entity = existing
                entity_type = "employee"
                approval.request_id = existing.id

        elif approval.target_table == "fleet":
            # Create new vehicle
            fleet_id = (
                form_data.get("registration")
                or scanned_data.get("employee_id")
                or f"TEMP{approval.id:04d}"
            )
            existing = db_session.query(Vehicle).filter_by(fleet_id=fleet_id).first()
            if not existing:
                new_vehicle = Vehicle(fleet_id=fleet_id, status="Active")
                db_session.add(new_vehicle)
                db_session.flush()

                # Generate QR code for new vehicle
                qr_data = f"VEH:{new_vehicle.id}:{new_vehicle.fleet_id}:{datetime.now().timestamp()}"
                qr_hash = hashlib.sha256(qr_data.encode()).hexdigest()[:32].upper()
                new_vehicle.qr_code = qr_hash

                new_entity = new_vehicle
                entity_type = "vehicle"
                approval.request_id = new_vehicle.id

        # Update gate log if exists
        gate_log = (
            db_session.query(GateLog)
            .filter_by(
                entity_id=approval.request_id, entity_name=approval.requester_name
            )
            .order_by(GateLog.scanned_at.desc())
            .first()
        )

        if gate_log:
            gate_log.access_granted = True
            gate_log.denial_reason = None
            if new_entity:
                gate_log.entity_id = new_entity.id
                if entity_type == "employee":
                    gate_log.employee_id = new_entity.id
                elif entity_type == "vehicle":
                    gate_log.vehicle_id = new_entity.id

            socketio.emit(
                "gate_scan",
                {
                    "type": entity_type or gate_log.access_type,
                    "name": approval.requester_name,
                    "direction": gate_log.direction,
                    "granted": True,
                    "reason": f"Approved - Added to {approval.target_table}",
                    "gate": gate_log.gate_location,
                    "time": datetime.now().strftime("%H:%M:%S"),
                },
            )

        db_session.commit()
        return jsonify(
            {
                "success": True,
                "message": f"Approved and added to {approval.target_table}",
                "entity_id": new_entity.id if new_entity else None,
                "entity_type": entity_type,
            }
        )
    return jsonify({"success": False, "message": "Approval not found"})


@app.route("/reject_request/<int:id>", methods=["POST"])
@login_required
@role_required(["admin", "manager"])
def reject_request(id):
    approval = db_session.query(Approval).filter_by(id=id).first()
    if approval:
        approval.status = "Rejected"
        approval.approved_by = session.get("username")
        approval.approval_date = _utcnow()
        approval.comments = request.json.get("comment", "")

        gate_log = (
            db_session.query(GateLog)
            .filter_by(
                entity_id=approval.request_id, entity_name=approval.requester_name
            )
            .order_by(GateLog.scanned_at.desc())
            .first()
        )

        if gate_log:
            gate_log.access_granted = False
            gate_log.denial_reason = (
                f"Rejected: {request.json.get('comment', 'No reason')}"
            )

            socketio.emit(
                "gate_scan",
                {
                    "type": gate_log.access_type,
                    "name": gate_log.entity_name,
                    "direction": gate_log.direction,
                    "granted": False,
                    "reason": f"Rejected: {request.json.get('comment', 'No reason')}",
                    "gate": gate_log.gate_location,
                    "time": datetime.now().strftime("%H:%M:%S"),
                },
            )

        db_session.commit()
        return jsonify({"success": True})
    return jsonify({"success": False, "message": "Approval not found"})


# ------------------- Barcode Generation (DataWedge/InfoWedge Compatible) -------------------


@app.route("/api/barcode/staging")
def generate_staging_barcode():
    """Generate Code 128 staging barcode image for DataWedge/InfoWedge device configuration."""
    try:
        host = request.host_url.rstrip("/")
        staging_data = f"DWCFG:{host}/api/config/datawedge"

        code128 = Code128(staging_data, writer=ImageWriter())
        buffer = io.BytesIO()
        code128.write(
            buffer,
            options={
                "module_height": 15.0,
                "module_width": 0.4,
                "quiet_zone": 6.5,
                "font_size": 8,
                "text_distance": 4.0,
                "write_text": True,
            },
        )
        buffer.seek(0)
        return send_file(buffer, mimetype="image/png")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/barcode/infowedge-setup")
def infowedge_setup_barcode():
    """Generate a QR code image that encodes InfoWedge IP output configuration.
    Scanning this on the C66 tells InfoWedge exactly where to send every scan.
    Format: IWCFG:<ip>:<port>:<path>
    """
    try:
        import qrcode as qrcode_lib

        host_ip = request.host.split(":")[0]
        port = request.host.split(":")[1] if ":" in request.host else "8080"
        # Encode the full target URL — InfoWedge IP output will POST raw barcode to this
        cfg = f"IWCFG:{host_ip}:{port}:/api/c66"
        img = qrcode_lib.make(cfg)
        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        return send_file(buffer, mimetype="image/png")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/config/infowedge-ip")
def infowedge_config():
    """Return InfoWedge IP output plugin configuration JSON.
    Tells the C66 InfoWedge app to redirect every scan to this server via TCP.
    """
    host_ip = request.host.split(":")[0]
    port_str = request.host.split(":")[1] if ":" in request.host else "8080"
    return jsonify(
        {
            "infowedge_config": {
                "profile_name": "MineGate",
                "ip_output": {
                    "enabled": True,
                    "address": host_ip,
                    "port": 9100,
                    "protocol": "TCP",
                    "data_format": "raw_barcode",
                },
                "intent_output": {
                    "enabled": True,
                    "action": "com.minegate.SCAN",
                    "extra_key": "barcodeData",
                    "delivery": "broadcast",
                },
                "http_output": {
                    "enabled": True,
                    "url": f"http://{host_ip}:{port_str}/api/c66",
                    "method": "POST",
                    "content_type": "text/plain",
                    "data": "{SCAN_BARCODE}",
                },
                "instructions": [
                    "Open InfoWedge on C66",
                    "Go to IP Output → Enable → set Address to "
                    + host_ip
                    + " Port 9100 Protocol TCP",
                    "OR: Go to Intent Output → Enable → Action: com.minegate.SCAN → Extra key: barcodeData → Broadcast",
                    "Every trigger pull will now send to MineGate",
                ],
            }
        }
    )


@app.route("/api/config/datawedge")
def datawedge_config():
    """Serve DataWedge/InfoWedge configuration in JSON format."""
    host = request.host_url.rstrip("/")

    config = {
        "configuration": {
            "profile_name": "MineGateScan",
            "config_mode": "CREATE_IF_NOT_EXIST",
            "plugin_settings": [
                {
                    "plugin_name": "BARCODE",
                    "reset_config": True,
                    "param_list": {
                        "scanner_selection": "auto",
                        "decoder_code128": "true",
                        "decoder_code39": "true",
                        "decoder_ean13": "true",
                        "decoder_qrcode": "true",
                    },
                },
                {
                    "plugin_name": "INTENT",
                    "reset_config": True,
                    "param_list": {
                        "intent_output_enabled": "true",
                        "intent_action": "com.minegate.SCAN",
                        "intent_category": "android.intent.category.DEFAULT",
                        "intent_delivery": "2",
                    },
                },
                {
                    "plugin_name": "HTTP",
                    "reset_config": True,
                    "param_list": {
                        "http_output_enabled": "true",
                        "http_url": f"{host}/api/scan_alt",
                        "http_method": "POST",
                        "http_content_type": "application/json",
                        "http_data": '{"barcode":{SCAN_BARCODE},"scanner":{SCANNER_NAME},"timestamp":{TIMESTAMP}}',
                    },
                },
            ],
            "barcode_input_enabled": True,
            "keystroke_output_enabled": False,
        }
    }

    return jsonify(config)


@app.route("/api/datawedge/staging")
def datawedge_staging_barcode():
    """Generate DataWedge staging barcode data (short format for Code 128)."""
    host = request.host_url.rstrip("/")
    # DataWedge staging format: DWCONFIG:<URL>
    # This tells DataWedge to fetch config from the URL
    staging_data = f"DWCFG:{host}/api/config/datawedge"

    return jsonify(
        {
            "staging_data": staging_data,
            "barcode_url": f"{host}/api/barcode/staging",
            "config_url": f"{host}/api/config/datawedge",
            "format": "Code128",
            "instructions": "Scan this barcode with your Zebra/Honeywell device to auto-configure InfoWedge",
        }
    )


@app.route("/api/app/download")
def download_apk():
    """Serve the MineGate Scanner APK for direct device installation."""
    import os

    apk_candidates = [
        os.path.join(os.path.dirname(__file__), "MineGateScanner.apk"),
        os.path.join(os.path.dirname(__file__), "QrMobile", "MineGateScanner.apk"),
        os.path.join(
            os.path.dirname(__file__),
            "QrMobile",
            "android",
            "app",
            "build",
            "outputs",
            "apk",
            "release",
            "app-release.apk",
        ),
        os.path.join(
            os.path.dirname(__file__),
            "QrMobile",
            "android",
            "app",
            "build",
            "outputs",
            "apk",
            "debug",
            "app-debug.apk",
        ),
    ]
    for path in apk_candidates:
        if os.path.exists(path):
            return send_file(
                path,
                mimetype="application/vnd.android.package-archive",
                as_attachment=True,
                download_name="MineGateScanner.apk",
            )
    return jsonify(
        {
            "error": "APK not built yet",
            "hint": "Run: cd QrMobile && npx expo run:android --variant release",
        }
    ), 404


@app.route("/api/barcode/app-download")
def app_download_barcode():
    """Generate Code 128 barcode encoding the APK direct-download URL.
    Scanning this barcode on any Android browser will download and install the app.
    """
    try:
        host = request.host_url.rstrip("/")
        download_url = f"{host}/api/app/download"
        code128 = Code128(download_url, writer=ImageWriter())
        buffer = io.BytesIO()
        code128.write(
            buffer,
            options={
                "module_height": 15.0,
                "module_width": 0.35,
                "quiet_zone": 6.5,
                "font_size": 7,
                "text_distance": 3.5,
                "write_text": True,
            },
        )
        buffer.seek(0)
        return send_file(buffer, mimetype="image/png")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/app/info")
def app_info():
    """Return app download info including barcode URL and QR code URL."""
    import os

    host = request.host_url.rstrip("/")
    apk_candidates = [
        os.path.join(os.path.dirname(__file__), "MineGateScanner.apk"),
        os.path.join(os.path.dirname(__file__), "QrMobile", "MineGateScanner.apk"),
        os.path.join(
            os.path.dirname(__file__),
            "QrMobile",
            "android",
            "app",
            "build",
            "outputs",
            "apk",
            "release",
            "app-release.apk",
        ),
        os.path.join(
            os.path.dirname(__file__),
            "QrMobile",
            "android",
            "app",
            "build",
            "outputs",
            "apk",
            "debug",
            "app-debug.apk",
        ),
    ]
    apk_ready = any(os.path.exists(p) for p in apk_candidates)
    return jsonify(
        {
            "app_name": "MineGate Scanner",
            "version": "1.2.0",
            "package": "com.minegate.scanner",
            "apk_ready": apk_ready,
            "download_url": f"{host}/api/app/download",
            "barcode_url": f"{host}/api/barcode/app-download",
            "config_barcode_url": f"{host}/api/barcode/staging",
            "datawedge_config_url": f"{host}/api/config/datawedge",
            "instructions": [
                "1. Scan the app-download barcode with any camera/browser on the C66",
                "2. Android will download and prompt to install MineGateScanner.apk",
                "3. After install, scan the config barcode to auto-configure the server IP",
                "4. The C66 hardware trigger will then send scans to this server",
            ],
        }
    )


# ------------------- Gate Scanner and Logs -------------------
@app.route("/gate_logs")
@login_required
@role_required(["admin", "security"])
def gate_logs():
    access_type = request.args.get("type", "")
    direction = request.args.get("direction", "")
    status = request.args.get("status", "")
    gate = request.args.get("gate", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    per_page = min(per_page, 200)  # cap

    # PERFORMANCE: noload('*') prevents lazy-loading of relationships
    # (employee, vehicle, visitor, equipment) during template iteration.
    from sqlalchemy.orm import noload as _noload

    query = (
        db_session.query(GateLog)
        .options(_noload("*"))
        .order_by(GateLog.scanned_at.desc())
    )

    if access_type:
        query = query.filter(GateLog.access_type == access_type)
    if direction:
        query = query.filter(GateLog.direction == direction)
    if status == "granted":
        query = query.filter(GateLog.access_granted.is_(True))
    elif status == "denied":
        query = query.filter(GateLog.access_granted.is_(False))
    if gate:
        query = query.filter(GateLog.gate_location == gate)
    if date_from:
        query = query.filter(
            GateLog.scanned_at >= datetime.strptime(date_from, "%Y-%m-%d")
        )
    if date_to:
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        query = query.filter(
            GateLog.scanned_at <= end_date.replace(hour=23, minute=59, second=59)
        )

    total = query.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    logs = query.offset((page - 1) * per_page).limit(per_page).all()

    return render_template(
        "gate_logs.html",
        logs=logs,
        selected_type=access_type,
        selected_direction=direction,
        selected_status=status,
        selected_gate=gate,
        selected_date_from=date_from,
        selected_date_to=date_to,
        page=page,
        total_pages=total_pages,
        total=total,
        per_page=per_page,
    )


def decode_qr_data(raw_data):
    """Universal QR decoder: parse any QR format into a normalized dict.

    Supports: JSON, pipe-delimited, CSV, URL query string, vCard, and
    'Key: Value' line-based formats. Returns dict with keys like
    employee_id, name, position, department, company, area, raw_data, format.
    """
    import re
    from urllib.parse import parse_qs

    result = {"raw_data": raw_data, "format": "unknown"}
    if not raw_data or not raw_data.strip():
        return result
    data = raw_data.strip()

    # 1. JSON
    if data.startswith("{"):
        try:
            parsed = json.loads(data)
            if isinstance(parsed, dict):
                result["format"] = "json"
                # Employee fields
                result["employee_id"] = (
                    parsed.get("employee_id")
                    or parsed.get("emp_code")
                    or parsed.get("id")
                )
                result["name"] = parsed.get("name") or parsed.get("full_name")
                result["position"] = (
                    parsed.get("position")
                    or parsed.get("job")
                    or parsed.get("job_title")
                )
                result["department"] = (
                    parsed.get("department")
                    or parsed.get("coy")
                    or parsed.get("company")
                )
                result["area"] = parsed.get("area")
                # Vehicle fields
                result["fleet_id"] = (
                    parsed.get("fleet_id")
                    or parsed.get("vehicle_id")
                    or parsed.get("fleet")
                    or parsed.get("registration")
                )
                result["vehicle_type"] = (
                    parsed.get("vehicle_type")
                    or parsed.get("type")
                    or parsed.get("model")
                )
                return result
        except (json.JSONDecodeError, ValueError):
            pass

    # 2. URL query string  (?id=123&name=John or full URL)
    if "?" in data and "=" in data:
        try:
            query_str = data.split("?", 1)[1] if "?" in data else data
            params = parse_qs(query_str, keep_blank_values=True)
            if params:
                result["format"] = "url_query"
                # Employee fields
                result["employee_id"] = (
                    params.get("id")
                    or params.get("emp_code")
                    or params.get("employee_id")
                    or [None]
                )[0]
                result["name"] = (
                    params.get("name") or params.get("full_name") or [None]
                )[0]
                result["position"] = (
                    params.get("position") or params.get("job") or [None]
                )[0]
                result["department"] = (
                    params.get("department")
                    or params.get("company")
                    or params.get("coy")
                    or [None]
                )[0]
                # Vehicle fields
                result["fleet_id"] = (
                    params.get("fleet_id")
                    or params.get("vehicle_id")
                    or params.get("fleet")
                    or params.get("registration")
                    or [None]
                )[0]
                result["vehicle_type"] = (
                    params.get("vehicle_type")
                    or params.get("type")
                    or params.get("model")
                    or [None]
                )[0]
                return result
        except Exception:
            pass

    # 3. vCard
    if data.upper().startswith("BEGIN:VCARD"):
        result["format"] = "vcard"
        fn_match = re.search(r"FN:(.*)", data)
        if fn_match:
            result["name"] = fn_match.group(1).strip()
        n_match = re.search(r"(?:^|\n)N:([^;]*);([^;]*)", data)
        if n_match:
            result["name"] = (
                result.get("name")
                or f"{n_match.group(2).strip()} {n_match.group(1).strip()}"
            )
        org_match = re.search(r"ORG:(.*)", data)
        if org_match:
            result["department"] = org_match.group(1).strip()
        title_match = re.search(r"TITLE:(.*)", data)
        if title_match:
            result["position"] = title_match.group(1).strip()
        return result

    # 4. Key: Value per line (e.g. "ID: 123\nName: John\nJob: Miner")
    kv_patterns = {
        "employee_id": r"(?:ID|Emp(?:loyee)?\s*(?:ID|Code))[:\s]+([^\|\n]+)",
        "name": r"(?:Name(?:\s+and\s+Surname)?)[:\s]+([^\|\n]+)",
        "position": r"(?:Job(?:\s*Title)?|Position|Occupation)[:\s]+([^\|\n]+)",
        "department": r"(?:Coy|Company|Dept|Department)[:\s]+([^\|\n]+)",
        "area": r"(?:Area|Section|Zone)[:\s]+([^\|\n]+)",
        "fleet_id": r"(?:Fleet(?:\s*ID)?|Vehicle\s*ID|Registration)[:\s]+([^\|\n]+)",
        "vehicle_type": r"(?:Vehicle\s*Type|Type|Model)[:\s]+([^\|\n]+)",
    }
    kv_found = False
    for key, pattern in kv_patterns.items():
        match = re.search(pattern, data, re.IGNORECASE)
        if match:
            result[key] = match.group(1).strip()
            kv_found = True
    if kv_found:
        result["format"] = "key_value"
        return result

    # 5. Pipe-delimited (e.g. "123|John Doe|Miner|Acme Corp" or "TRUCK001|Volvo|Dump Truck")
    if "|" in data:
        parts = [p.strip() for p in data.split("|") if p.strip()]
        if len(parts) >= 2:
            result["format"] = "pipe"
            # Check if first part looks like a vehicle ID (letters/numbers mix, often starts with letters)
            first_part = parts[0]
            is_vehicle_id = bool(
                re.match(r"^[A-Z]{2,}\d+", first_part, re.IGNORECASE)
            ) or any(
                x in first_part.upper() for x in ["TRUCK", "LDV", "DUMP", "EXCAVATOR"]
            )

            if is_vehicle_id:
                result["fleet_id"] = first_part
                result["vehicle_type"] = parts[1] if len(parts) > 1 else None
            else:
                # Assume employee format
                result["employee_id"] = (
                    first_part
                    if first_part.replace("-", "").replace("_", "").isalnum()
                    else None
                )
                result["name"] = parts[1] if len(parts) > 1 else None
                result["position"] = parts[2] if len(parts) > 2 else None
                result["department"] = parts[3] if len(parts) > 3 else None
                result["area"] = parts[4] if len(parts) > 4 else None
            return result

    # 6. CSV (e.g. "123,John Doe,Miner,Acme Corp" or "TRUCK001,Volvo,2020")
    if "," in data and "\n" not in data:
        parts = [p.strip().strip('"') for p in data.split(",") if p.strip()]
        if len(parts) >= 2:
            result["format"] = "csv"
            first_part = parts[0]
            # Check if looks like vehicle ID
            is_vehicle_id = bool(
                re.match(r"^[A-Z]{2,}\d+", first_part, re.IGNORECASE)
            ) or any(
                x in first_part.upper() for x in ["TRUCK", "LDV", "DUMP", "EXCAVATOR"]
            )

            if is_vehicle_id:
                result["fleet_id"] = first_part
                result["vehicle_type"] = parts[1] if len(parts) > 1 else None
            else:
                # Assume employee format
                result["employee_id"] = (
                    first_part
                    if first_part.replace("-", "").replace("_", "").isalnum()
                    else None
                )
                result["name"] = parts[1] if len(parts) > 1 else None
                result["position"] = parts[2] if len(parts) > 2 else None
                result["department"] = parts[3] if len(parts) > 3 else None
            return result

    # 7. Plain text fallback — treat entire string as an ID or name
    result["format"] = "plain"
    if data.replace("-", "").replace("_", "").isalnum() and len(data) <= 50:
        result["employee_id"] = data
    else:
        result["name"] = data[:100]
    return result


def _get_gate_name_from_ip(ip_address, scanned_by, default_gate_location=None):
    """Look up gate name from IP address in gate_mappings table.

    Args:
        ip_address: The IP address of the scanner
        scanned_by: The scanner identifier string (e.g., "infowedge:192.168.0.160:9100")
        default_gate_location: Fallback gate location if no mapping found

    Returns:
        Gate name string (e.g., "Extension Gate 1") or default/fallback
    """
    if not ip_address and not scanned_by:
        return default_gate_location or "Main Gate"

    # Try to extract IP from scanned_by if ip_address is empty
    lookup_ip = ip_address
    if not lookup_ip and scanned_by:
        # Extract IP from formats like "infowedge:192.168.0.160:9100"
        import re

        ip_match = re.search(r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", scanned_by)
        if ip_match:
            lookup_ip = ip_match.group(1)

    if not lookup_ip:
        return default_gate_location or "Main Gate"

    # Look up in gate_mappings table
    try:
        mapping = (
            db_session.query(GateMapping)
            .filter(GateMapping.ip_address == lookup_ip, GateMapping.is_active)
            .first()
        )

        if mapping:
            return mapping.gate_name
    except Exception:
        # If table doesn't exist yet or other error, return default
        pass

    return default_gate_location or f"Gate-{lookup_ip}"


def _process_qr_scan(
    qr_hash, direction, gate_location, scanned_by, ip_address, user_agent
):
    """Process a QR code scan and return entity info and access decision.

    Thin wrapper around services.scan_service.process_qr_scan().
    """
    from services.scan_service import process_qr_scan

    return process_qr_scan(
        qr_hash,
        direction,
        gate_location,
        scanned_by,
        ip_address,
        user_agent,
        socketio,
    )


@app.route("/kiosk")
def kiosk_scanner():
    """Full-screen kiosk page for C66 keyboard emulator and InfoWedge browser mode.
    Open this URL on the C66's browser / WebView — it captures all barcode input
    and shows a forced full-screen GREEN/RED overlay.
    """
    return render_template("kiosk_scanner.html")


_LOCAL_PREFIXES = ("192.168.", "10.", "172.", "127.")


def _is_local_ip(ip):
    """Return True if the IP belongs to a RFC-1918 / loopback range."""
    return any(ip.startswith(p) for p in _LOCAL_PREFIXES)


@app.route("/api/c66", methods=["POST", "GET"])
def c66_ingest():
    """Dedicated C66/InfoWedge ingest endpoint.
    Accepts: plain text body, JSON, or ?data= query param.
    No API key required — restricted to local network IPs.
    InfoWedge IP Output: set URL to http://<server>:8080/api/c66
    """
    ip_address = request.remote_addr
    user_agent = request.headers.get("User-Agent", "InfoWedge/C66")

    # Accept from local IPs only
    local_prefixes = ("192.168.", "10.", "172.", "127.")
    if not any(ip_address.startswith(p) for p in local_prefixes):
        return jsonify({"success": False, "message": "Local network only"}), 403

    content_type = request.content_type or ""
    qr_hash = None

    if "application/json" in content_type:
        data = request.get_json(silent=True) or {}
        qr_hash = (
            data.get("barcodeData")
            or data.get("barcode")
            or data.get("qr_code")
            or data.get("data")
            or data.get("barcodeStringData")
            or data.get("scanData")
        )
    elif request.args.get("data"):
        qr_hash = request.args.get("data")
    else:
        raw = request.get_data(as_text=True).strip()
        if raw:
            qr_hash = raw

    if not qr_hash:
        return "OK", 200  # InfoWedge sometimes sends empty keepalives

    qr_hash = qr_hash.strip()
    result = _process_qr_scan(
        qr_hash, "AUTO", "C66 Gate", f"c66:{ip_address}", ip_address, user_agent
    )
    _ensure_device_exists(ip_address)

    # Log scan decisions for log_viewer dashboard
    found_in = "none"
    if result["entity_type"] == "employee":
        found_in = "employees"
    elif result["entity_type"] == "vehicle":
        found_in = "vehicles"
    elif result["entity_type"] == "visitor":
        found_in = "visitors"

    if not result["access_granted"] and result["denial_reason"]:
        pending = (
            db_session.query(Approval)
            .filter(
                Approval.status == "Pending", Approval.scanned_data.contains(qr_hash)
            )
            .first()
        )
        if pending:
            found_in = "pending"

    scan_log_data = {
        "code": qr_hash,
        "foundIn": found_in,
        "granted": result["access_granted"],
        "entity": result["entity_name"] or "Unknown",
        "direction": "AUTO",
        "type": result["entity_type"] or "QR",
    }
    print(f"SCAN LOG: {json.dumps(scan_log_data)}", flush=True)

    entity_name = result.get("entity_name") or "Unknown"
    granted = result["access_granted"]
    denial = result.get("denial_reason")

    tts_msg = (
        f"Access granted. {entity_name}"
        if granted
        else f"Access denied. {denial or ''}"
    )
    if not granted and entity_name != "Unknown":
        tts_msg = f"Access denied for {entity_name}. {denial or ''}"

    try:
        socketio.emit(
            "scan_result",
            {
                "success": granted,
                "entity_name": entity_name,
                "message": denial or "",
                "scanner": ip_address,
                "protocol": "HTTP-C66",
                "notification": "approved" if granted else "denied",
                "tts_message": tts_msg,
            },
        )
    except Exception:
        pass

    return jsonify(
        {
            "success": granted,
            "open_gate": granted,
            "message": denial or ("Access granted" if granted else "Access denied"),
            "name": entity_name,
            "entity_name": entity_name,
            "entity_type": result.get("entity_type", ""),
            "status": "approved" if granted else "denied",
            "denial_reason": denial,
            "parsed_data": result.get("parsed_qr"),
            "is_unknown": entity_name == "Unknown" or "Unassigned" in str(entity_name),
            "notification": "approved" if granted else "denied",
            "tts_message": tts_msg,
        }
    )


@app.route("/message")
def message_endpoint():
    """InfoWedge connectivity check endpoint.

    InfoWedge makes GET requests to this endpoint to verify server connectivity.
    Auto-registers device on first check-in.
    """
    client_ip = request.remote_addr
    device_name = request.args.get("device", "Unknown")
    app_name = request.args.get("app", "")

    _ensure_device_exists(client_ip)
    print(f"INFOWEDGE CHECK-IN: device={device_name} app={app_name} from {client_ip}")

    return "OK", 200, {"Content-Type": "text/plain"}


# ═══════════════════════════════════════════════════════════════════════
# RFID SCAN ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════


@app.route("/api/rfid_ingest", methods=["POST", "GET"])
def rfid_ingest():
    """Raw RFID TCP ingest endpoint for direct scanner connections.

    Accepts raw RFID tag data from TCP stream scanners.
    IP whitelisting recommended for 192.168.0.187 (RFID scanner)
    """
    ip_address = request.remote_addr
    user_agent = request.headers.get("User-Agent", "RFID Reader")

    # Accept from local IPs and the configured RFID scanner
    allowed_ips = ("192.168.", "10.", "172.", "127.", "192.168.0.187")
    if (
        not any(ip_address.startswith(p) for p in allowed_ips)
        and ip_address != "192.168.0.187"
    ):
        return jsonify(
            {"success": False, "message": "Local network or RFID scanner only"}
        ), 403

    content_type = request.content_type or ""
    rfid_tag = None

    if "application/json" in content_type:
        data = request.get_json(silent=True) or {}
        rfid_tag = (
            data.get("rfid_tag")
            or data.get("tag")
            or data.get("epc")
            or data.get("data")
            or data.get("uid")
        )
    elif request.args.get("tag"):
        rfid_tag = request.args.get("tag")
    elif request.args.get("rfid"):
        rfid_tag = request.args.get("rfid")
    else:
        raw = request.get_data(as_text=True).strip()
        if raw:
            rfid_tag = raw

    if not rfid_tag:
        return "OK", 200  # Keep-alive response

    # Format and process
    formatted_tag = _format_rfid_tag(rfid_tag.strip())
    result = _process_rfid_scan(
        formatted_tag, "AUTO", "RFID Gate", f"rfid:{ip_address}", ip_address, user_agent
    )

    # Emit for real-time monitoring
    try:
        socketio.emit(
            "rfid_scan_result",
            {
                "success": result["access_granted"],
                "entity_name": result.get("entity_name", ""),
                "message": result.get("denial_reason", ""),
                "rfid_tag": formatted_tag,
                "scanner": ip_address,
                "protocol": "RFID-RAW",
            },
        )
    except Exception:
        pass

    return jsonify(
        {
            "success": result["access_granted"],
            "open_gate": result["access_granted"],
            "message": result["denial_reason"]
            or ("Access granted" if result["access_granted"] else "Access denied"),
            "name": result.get("entity_name") or "Unknown",
            "entity_name": result.get("entity_name", ""),
            "entity_type": result.get("entity_type", ""),
            "rfid_tag": formatted_tag,
            "status": "approved" if result["access_granted"] else "denied",
        }
    )


def _format_rfid_tag(raw_tag):
    """Format and normalize RFID tag data from various formats.

    Supports:
    - EPC Gen2 (96-bit): E20034150200108022001F6D
    - ISO 14443A (MIFARE): 04:A2:3B:1C or 04A23B1C
    - Raw hex with/without separators
    """
    if not raw_tag:
        return None

    # Remove common separators and whitespace
    tag = raw_tag.strip().upper()
    tag = tag.replace(":", "").replace("-", "").replace(" ", "").replace(".", "")

    # Remove any prefixes some readers add
    prefixes_to_strip = ["EPC:", "UID:", "TAG:", "RFID:", "[", "]"]
    for prefix in prefixes_to_strip:
        tag = tag.replace(prefix, "")

    # Validate hex content
    if not all(c in "0123456789ABCDEF" for c in tag):
        # Non-hex tag, keep original but normalized
        return tag

    return tag


def _process_rfid_scan(
    rfid_tag, direction, gate_location, scanned_by, ip_address, user_agent
):
    """Process an RFID tag scan and return entity info and access decision.

    Similar to _process_qr_scan but looks up by rfid_tag field.
    """
    entity = None
    entity_type = None
    entity_id = None
    entity_name = None
    access_granted = False
    denial_reason = None

    # Try to find entity by RFID tag
    employee = db_session.query(Employee).filter_by(rfid_tag=rfid_tag).first()

    if employee:
        entity = employee
        entity_type = "employee"
        entity_id = employee.id
        entity_name = f"{employee.first_name} {employee.surname}"

    if not entity:
        vehicle = db_session.query(Vehicle).filter_by(rfid_tag=rfid_tag).first()
        if vehicle:
            entity = vehicle
            entity_type = "vehicle"
            entity_id = vehicle.id
            entity_name = vehicle.fleet_id

    if not entity:
        visitor = db_session.query(Visitor).filter_by(rfid_tag=rfid_tag).first()
        if visitor:
            entity = visitor
            entity_type = "visitor"
            entity_id = visitor.id
            entity_name = visitor.name

    if not entity:
        equipment = db_session.query(Equipment).filter_by(rfid_tag=rfid_tag).first()
        if equipment:
            entity = equipment
            entity_type = "equipment"
            entity_id = equipment.id
            entity_name = equipment.radio_id

    # Auto-direction logic (same as QR scan)
    if entity_id and entity_type:
        # PERFORMANCE: noload('*') prevents lazy-loading relationships
        # when we only need the direction column.
        from sqlalchemy.orm import noload as _noload

        last_log = (
            db_session.query(GateLog)
            .options(_noload("*"))
            .filter(
                GateLog.entity_id == entity_id,
                GateLog.access_type == entity_type,
                GateLog.access_granted,
            )
            .order_by(GateLog.scanned_at.desc())
            .first()
        )
        if last_log and last_log.direction == "IN":
            direction = "OUT"
        else:
            direction = "IN"
    else:
        direction = "IN"
        entity_name = "Unknown"
        entity_type = "unknown"

    # Access decision logic
    if entity:
        if entity_type == "employee":
            if entity.status != "Active":
                access_granted = False
                denial_reason = f"Employee status is {entity.status}"
            else:
                access_granted = True
        elif entity_type == "vehicle":
            if entity.status != "Active":
                access_granted = False
                denial_reason = f"Vehicle status is {entity.status}"
            else:
                access_granted = True
        elif entity_type == "visitor":
            if entity.status != "Checked In":
                access_granted = False
                denial_reason = f"Visitor status is {entity.status}"
            else:
                access_granted = True
        elif entity_type == "equipment":
            if entity.status != "Active":
                access_granted = False
                denial_reason = f"Equipment status is {entity.status}"
            else:
                access_granted = True
    else:
        access_granted = False
        denial_reason = "RFID tag not registered"

    # Create gate log entry
    gate_log = GateLog(
        access_type=entity_type or "unknown",
        entity_id=entity_id,
        entity_name=entity_name,
        direction=direction,
        qr_data=rfid_tag,  # Store RFID in qr_data field for compatibility
        access_granted=access_granted,
        denial_reason=denial_reason,
        gate_location=gate_location,
        scanned_by=scanned_by,
        ip_address=ip_address,
        user_agent=user_agent,
    )
    db_session.add(gate_log)
    db_session.commit()

    # Invalidate caches since gate log data changed
    from routes.dashboard import invalidate_dashboard_cache
    from routes.monitoring import invalidate_monitoring_cache
    invalidate_dashboard_cache()
    invalidate_monitoring_cache()

    return {
        "access_granted": access_granted,
        "denial_reason": denial_reason,
        "entity_type": entity_type,
        "entity_name": entity_name,
        "entity_id": entity_id,
        "direction": direction,
        "rfid_tag": rfid_tag,
    }


@app.route("/api/verify-visitor", methods=["POST"])
def verify_visitor_mobile():
    """Mobile app visitor check-in/check-out endpoint."""
    data = request.get_json()
    qr_data = data.get("qr_data") or data.get("code", "")
    device_info = data.get("device_info", "Mobile Scanner")
    ip_address = request.remote_addr
    user_agent = request.headers.get("User-Agent", "")

    # Try to extract visitor ID from various QR formats
    visitor_id = None

    # Parse JSON format: {"type": "visitor", "id": 123, "name": "..."}
    try:
        qr_json = json.loads(qr_data)
        if qr_json.get("type") == "visitor":
            visitor_id = qr_json.get("id")
            qr_json.get("name")
    except (json.JSONDecodeError, AttributeError):
        pass

    # Look up visitor by QR code hash or ID
    visitor = None
    if visitor_id:
        visitor = db_session.query(Visitor).filter_by(id=visitor_id).first()
    if not visitor and qr_data:
        visitor = db_session.query(Visitor).filter_by(qr_code=qr_data).first()

    if not visitor:
        return jsonify(
            {
                "valid": False,
                "type": "error",
                "message": "Visitor not found",
                "visitor": None,
            }
        ), 404

    # Determine action based on current status
    is_checkin = visitor.status != "Checked In"

    if is_checkin:
        # Check in
        visitor.status = "Checked In"
        visitor.check_in_time = _utcnow()
        direction = "IN"
        result_type = "check_in"
        message = f"Welcome {visitor.name}! Check-in recorded."
    else:
        # Check out
        visitor.status = "Checked Out"
        visitor.check_out_time = _utcnow()
        direction = "OUT"
        result_type = "check_out"

        # Calculate duration
        duration = "Unknown"
        if visitor.check_in_time:
            diff = visitor.check_out_time - visitor.check_in_time
            hours = int(diff.total_seconds() // 3600)
            minutes = int((diff.total_seconds() % 3600) // 60)
            duration = f"{hours}h {minutes}m"
        message = f"Goodbye {visitor.name}! Duration: {duration}"

    db_session.commit()

    # Log the scan
    gate_log = GateLog(
        access_type="visitor",
        entity_id=visitor.id,
        entity_name=visitor.name,
        direction=direction,
        qr_data=qr_data,
        access_granted=True,
        denial_reason=None,
        gate_location=device_info,
        scanned_by=f"mobile-{ip_address}",
        ip_address=ip_address,
        user_agent=user_agent,
        visitor_id=visitor.id,
    )
    db_session.add(gate_log)
    db_session.commit()

    # Invalidate caches since gate log data changed
    from routes.dashboard import invalidate_dashboard_cache
    from routes.monitoring import invalidate_monitoring_cache
    invalidate_dashboard_cache()
    invalidate_monitoring_cache()

    # Emit to dashboard
    socketio.emit(
        "gate_scan",
        {
            "type": "visitor",
            "name": visitor.name,
            "direction": direction,
            "granted": True,
            "reason": message,
            "gate": device_info,
            "time": datetime.now().strftime("%H:%M:%S"),
        },
    )

    return jsonify(
        {
            "valid": True,
            "type": result_type,
            "message": message,
            "duration": duration if not is_checkin else None,
            "visitor": {
                "id": visitor.id,
                "name": visitor.name,
                "company": visitor.company,
                "purpose": visitor.purpose,
                "status": visitor.status,
            },
        }
    )


@app.route("/api/gate_logs")
@login_required
def api_gate_logs():
    limit = request.args.get("limit", 100, type=int)
    limit = min(limit, 1000)  # cap to prevent excessive memory use
    page = request.args.get("page", 1, type=int)
    page = max(1, page)

    # PERFORMANCE: noload('*') prevents lazy-loading of relationships.
    from sqlalchemy.orm import noload as _noload

    logs = (
        db_session.query(GateLog)
        .options(_noload("*"))
        .order_by(GateLog.scanned_at.desc())
        .offset((page - 1) * limit)
        .limit(limit)
        .all()
    )
    total = db_session.query(GateLog).count()
    return jsonify(
        {
            "logs": [
                {
                    "id": log.id,
                    "type": log.access_type,
                    "name": log.entity_name,
                    "direction": log.direction,
                    "granted": log.access_granted,
                    "reason": log.denial_reason,
                    "gate": log.gate_location,
                    "scanned_by": log.scanned_by,
                    "time": log.scanned_at.strftime("%Y-%m-%d %H:%M:%S"),
                }
                for log in logs
            ],
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": max(1, (total + limit - 1) // limit),
        }
    )


@app.route("/api/recent_activity")
@login_required
def api_recent_activity():
    try:
        today = datetime.now().date()
        today_start = datetime.combine(today, datetime.min.time())
        # PERFORMANCE: noload('*') prevents lazy-loading relationships.
        from sqlalchemy.orm import noload as _noload

        logs = (
            db_session.query(GateLog)
            .options(_noload("*"))
            .filter(GateLog.scanned_at >= today_start)
            .order_by(GateLog.scanned_at.desc())
            .limit(10)
            .all()
        )
        return jsonify(
            [
                {
                    "type": log.access_type,
                    "name": log.entity_name,
                    "direction": log.direction,
                    "granted": log.access_granted,
                    "reason": log.denial_reason,
                    "gate": log.gate_location,
                    "time": log.scanned_at.strftime("%H:%M:%S"),
                }
                for log in logs
            ]
        )
    except Exception:
        return jsonify([])


@app.route("/health", methods=["GET"])
def health_check():
    """Health check endpoint for mobile apps."""
    return jsonify(
        {
            "status": "ok",
            "service": "mine-management-api",
            "timestamp": datetime.now().isoformat(),
        }
    )


@app.route("/export/gate_logs/excel")
@login_required
@role_required(["admin"])
def export_gate_logs_excel():
    access_type = request.args.get("type", "")
    direction = request.args.get("direction", "")
    status = request.args.get("status", "")
    gate = request.args.get("gate", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    # PERFORMANCE: noload('*') prevents lazy-loading of relationships.
    # NOTE: Streaming opportunity — for very large datasets, replace .all()
    # with yield_per() + server-side cursor to avoid loading all 50k rows
    # into memory at once.
    from sqlalchemy.orm import noload as _noload

    query = db_session.query(GateLog).order_by(GateLog.scanned_at.desc())

    if access_type:
        query = query.filter(GateLog.access_type == access_type)
    if direction:
        query = query.filter(GateLog.direction == direction)
    if status == "granted":
        query = query.filter(GateLog.access_granted.is_(True))
    elif status == "denied":
        query = query.filter(GateLog.access_granted.is_(False))
    if gate:
        query = query.filter(GateLog.gate_location == gate)
    if date_from:
        query = query.filter(
            GateLog.scanned_at >= datetime.strptime(date_from, "%Y-%m-%d")
        )
    if date_to:
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        query = query.filter(
            GateLog.scanned_at <= end_date.replace(hour=23, minute=59, second=59)
        )

    # Use write_only mode to optimize memory usage (P0 performance fix)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="Gate Logs")
    headers = [
        "ID",
        "Type",
        "Name",
        "Direction",
        "Granted",
        "Reason",
        "Gate",
        "Scanned By",
        "Time",
    ]
    ws.append(headers)

    logs = query.options(_noload("*")).limit(50000).all()
    for log in logs:
        ws.append(
            [
                _sanitize_cell(v)
                for v in [
                    log.id,
                    log.access_type,
                    log.entity_name,
                    log.direction,
                    "Yes" if log.access_granted else "No",
                    log.denial_reason or "",
                    log.gate_location,
                    log.scanned_by,
                    log.scanned_at.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            ]
        )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="gate_logs.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ------------------- Export Routes -------------------
@app.route("/export/employees/excel")
@login_required
@role_required(["admin", "manager"])
def export_employees_excel():
    # NOTE: Streaming opportunity — for very large employee datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    employees = db_session.query(Employee).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "ID",
        "EmpCode",
        "Initials",
        "FirstName",
        "SecondName",
        "Surname",
        "ID Number",
        "JobTitle",
        "Induction",
        "Induction Expiry",
        "Medical",
        "Medical Expiry",
        "Status",
    ]
    ws.append(headers)
    for emp in employees:
        ws.append(
            [
                _sanitize_cell(v)
                for v in [
                    emp.id,
                    emp.emp_code,
                    emp.initials or "",
                    emp.first_name,
                    emp.second_name or "",
                    emp.surname,
                    emp.id_number,
                    emp.job_title or "",
                    emp.induction or "",
                    emp.induction_expiry.strftime("%Y-%m-%d")
                    if emp.induction_expiry
                    else "",
                    emp.medical or "",
                    emp.medical_expiry.strftime("%Y-%m-%d")
                    if emp.medical_expiry
                    else "",
                    emp.status,
                ]
            ]
        )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="employees_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/visitors/excel")
@login_required
@role_required(["admin", "manager"])
def export_visitors_excel():
    # PERFORMANCE: eager-load host relationship to avoid N+1 queries
    # when accessing visitor.host.name for each row.
    from sqlalchemy.orm import joinedload as _joinedload

    # NOTE: Streaming opportunity — for very large visitor datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    visitors = db_session.query(Visitor).options(_joinedload(Visitor.host)).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitors"
    headers = [
        "ID",
        "Name",
        "Company",
        "Purpose",
        "Host",
        "Check In",
        "Check Out",
        "Status",
    ]
    ws.append(headers)
    for visitor in visitors:
        ws.append(
            [
                _sanitize_cell(v)
                for v in [
                    visitor.id,
                    visitor.name,
                    visitor.company or "",
                    visitor.purpose or "",
                    visitor.host.name if visitor.host else "",
                    visitor.check_in_time.strftime("%Y-%m-%d %H:%M"),
                    visitor.check_out_time.strftime("%Y-%m-%d %H:%M")
                    if visitor.check_out_time
                    else "",
                    visitor.status,
                ]
            ]
        )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="visitors_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/fleet/excel")
@login_required
@role_required(["admin", "manager"])
def export_fleet_excel():
    # NOTE: Streaming opportunity — for very large fleet datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    vehicles = db_session.query(Vehicle).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fleet"
    ws.append(["Fleet ID"])
    for vehicle in vehicles:
        ws.append([vehicle.fleet_id])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="fleet_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/fleet/pdf")
@login_required
@role_required(["admin", "manager"])
def export_fleet_pdf():
    # NOTE: Streaming opportunity — for very large fleet datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    vehicles = db_session.query(Vehicle).all()

    headers = ["Fleet ID"]
    data = []
    for vehicle in vehicles:
        data.append([vehicle.fleet_id])

    output = generate_pdf("Fleet Report", headers, data, None)

    return send_file(
        output,
        as_attachment=True,
        download_name="fleet_export.pdf",
        mimetype="application/pdf",
    )


@app.route("/export/equipment/excel")
@login_required
@role_required(["admin", "manager"])
def export_equipment_excel():
    # NOTE: Streaming opportunity — for very large equipment datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    items = db_session.query(Equipment).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment"
    ws.append(["Radio ID", "Status", "Created At"])
    for item in items:
        ws.append([item.radio_id, item.status, item.created_at.strftime("%Y-%m-%d")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="equipment_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/equipment/pdf")
@login_required
@role_required(["admin", "manager"])
def export_equipment_pdf():
    # NOTE: Streaming opportunity — for very large equipment datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    items = db_session.query(Equipment).all()

    headers = ["Radio ID", "Status", "Registered On"]
    data = []
    for item in items:
        data.append([item.radio_id, item.status, item.created_at.strftime("%Y-%m-%d")])

    output = generate_pdf("Equipment Report", headers, data, None)

    return send_file(
        output,
        as_attachment=True,
        download_name="equipment_export.pdf",
        mimetype="application/pdf",
    )


def _get_base_url():
    """Get the base URL for the system, preferring the configured SiteSetting."""
    setting = db_session.query(SiteSetting).filter_by(key="system_base_url").first()
    if setting and setting.value:
        return setting.value.rstrip("/")

    # Fallback to current request or local IP
    try:
        return request.host_url.rstrip("/")
    except Exception:
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            _ip = s.getsockname()[0]
            s.close()
            return f"http://{_ip}:8080"
        except Exception:
            return "http://localhost:8080"


@app.route("/export/equipment/qr-zip")
@login_required
@role_required(["admin", "manager"])
def export_equipment_qr_zip():
    """Export all equipment QR codes as a ZIP file."""
    import zipfile

    import qrcode

    items = db_session.query(Equipment).filter(Equipment.qr_code.isnot(None)).all()
    if not items:
        return "No QR codes found for equipment. Please generate QR codes first.", 404

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zf:
        for item in items:
            # Generate QR code if missing or use radio_id as fallback
            qr_hash = item.qr_code or item.radio_id
            qr_url = f"{_get_base_url()}/s/{qr_hash}"
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(qr_url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")

            img_buffer = io.BytesIO()
            img.save(img_buffer, format="PNG")
            zf.writestr(f"equipment_{item.radio_id}.png", img_buffer.getvalue())

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name="equipment_qr_codes.zip",
        mimetype="application/zip",
    )


@app.route("/export/employees/qr-zip")
@login_required
@role_required(["admin", "security"])
def export_employees_qr_zip():
    """Export all employee QR codes as a ZIP file."""
    import zipfile

    employees = db_session.query(Employee).filter(Employee.qr_code.isnot(None)).all()

    if not employees:
        return "No QR codes found for employees. Please generate QR codes first.", 404

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for emp in employees:
            qr_url = f"{_get_base_url()}/s/{emp.qr_code}"
            qr = qrcode.QRCode(version=4, box_size=20, border=4)
            qr.add_data(qr_url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")

            img_buffer = io.BytesIO()
            img.save(img_buffer, format="PNG")
            img_buffer.seek(0)

            filename = f"{emp.emp_code}_{emp.surname.replace(' ', '_')}_qr.png"
            zip_file.writestr(filename, img_buffer.getvalue())

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name="employees_qr_codes.zip",
        mimetype="application/zip",
    )


@app.route("/export/fleet/qr-zip")
@login_required
@role_required(["admin", "security"])
def export_fleet_qr_zip():
    """Export all vehicle QR codes as a ZIP file."""
    import zipfile

    vehicles = db_session.query(Vehicle).filter(Vehicle.qr_code.isnot(None)).all()

    if not vehicles:
        return "No QR codes found for vehicles. Please generate QR codes first.", 404

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for vehicle in vehicles:
            qr_url = f"{_get_base_url()}/s/{vehicle.qr_code}"
            qr = qrcode.QRCode(version=4, box_size=20, border=4)
            qr.add_data(qr_url)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white").convert(
                "RGB"
            )

            # Add text overlay with Fleet ID
            from PIL import Image as PilImage
            from PIL import ImageDraw, ImageFont

            try:
                font_large = ImageFont.truetype(
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20
                )
                font_small = ImageFont.truetype(
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16
                )
            except Exception:
                try:
                    font_large = ImageFont.truetype(
                        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                        20,
                    )
                    font_small = ImageFont.truetype(
                        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                        16,
                    )
                except Exception:
                    font_large = ImageFont.load_default()
                    font_small = ImageFont.load_default()

            text_padding = 65
            canvas = PilImage.new(
                "RGB", (qr_img.width, qr_img.height + text_padding), "white"
            )
            canvas.paste(qr_img, (0, 0))
            draw = ImageDraw.Draw(canvas)
            img_width = canvas.width

            # Draw Fleet ID label
            label_text = f"Fleet ID: {vehicle.fleet_id}"
            try:
                bbox = draw.textbbox((0, 0), label_text, font=font_large)
                text_width = bbox[2] - bbox[0]
            except AttributeError:
                text_width = len(label_text) * 10
            x = (img_width - text_width) // 2
            draw.text((x, qr_img.height + 8), label_text, fill="black", font=font_large)

            # Draw "Vehicle" label
            id_text = "Vehicle"
            try:
                bbox2 = draw.textbbox((0, 0), id_text, font=font_small)
                text_width2 = bbox2[2] - bbox2[0]
            except AttributeError:
                text_width2 = len(id_text) * 8
            x2 = (img_width - text_width2) // 2
            draw.text(
                (x2, qr_img.height + 34), id_text, fill="#444444", font=font_small
            )

            img_buffer = io.BytesIO()
            canvas.save(img_buffer, format="PNG")
            img_buffer.seek(0)

            filename = f"{vehicle.fleet_id.replace(' ', '_')}_qr.png"
            zip_file.writestr(filename, img_buffer.getvalue())

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name="fleet_qr_codes.zip",
        mimetype="application/zip",
    )


# ------------------- PDF Export Helper -------------------
class HeaderFooterCanvas:
    """Custom canvas for drawing header with logo and footer with page numbers."""

    def __init__(self, logo_path=None):
        self.logo_path = logo_path
        self.logo_available = logo_path and os.path.exists(logo_path)

    def draw_header(self, canvas, doc):
        """Draw header with logo and company name."""
        canvas.saveState()

        # Draw logo if available
        if self.logo_available:
            try:
                logo_width = 1.2 * inch
                logo_height = 0.8 * inch
                canvas.drawImage(
                    self.logo_path,
                    30,
                    doc.pagesize[1] - 50,
                    width=logo_width,
                    height=logo_height,
                    preserveAspectRatio=True,
                    anchor="nw",
                )
                text_x = 30 + logo_width + 15
            except Exception:
                text_x = 30
        else:
            text_x = 30

        # Company name and subtitle
        canvas.setFont("Helvetica-Bold", 16)
        canvas.setFillColor(colors.HexColor("#e10600"))
        canvas.drawString(text_x, doc.pagesize[1] - 35, "Arch-System")

        canvas.setFont("Helvetica", 10)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(
            text_x, doc.pagesize[1] - 50, "Professional Mining Operations Management"
        )

        # Header line
        canvas.setStrokeColor(colors.HexColor("#e10600"))
        canvas.setLineWidth(2)
        canvas.line(
            30, doc.pagesize[1] - 65, doc.pagesize[0] - 30, doc.pagesize[1] - 65
        )

        canvas.restoreState()

    def draw_footer(self, canvas, doc):
        """Draw footer with page numbers and disclaimer."""
        canvas.saveState()

        # Footer line
        canvas.setStrokeColor(colors.HexColor("#cccccc"))
        canvas.setLineWidth(0.5)
        canvas.line(30, 50, doc.pagesize[0] - 30, 50)

        # Page number
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.HexColor("#666666"))
        page_text = f"Page {doc.page}"
        canvas.drawCentredString(doc.pagesize[0] / 2, 35, page_text)

        # Disclaimer
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#999999"))
        disclaimer = "Confidential - Arch-System - Generated Report"
        canvas.drawCentredString(doc.pagesize[0] / 2, 22, disclaimer)

        # Timestamp on right
        timestamp = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        canvas.drawRightString(doc.pagesize[0] - 30, 35, timestamp)

        canvas.restoreState()


def generate_pdf(title, headers, data, filters=None):
    """Generate a professional styled PDF report using ReportLab."""
    output = io.BytesIO()

    # Setup document with proper margins for header/footer
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=90,  # Space for header
        bottomMargin=70,  # Space for footer
    )

    styles = getSampleStyleSheet()
    story = []

    # Custom styles
    report_title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=22,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=6,
        alignment=TA_LEFT,
        fontName="Helvetica-Bold",
    )

    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=12,
        textColor=colors.HexColor("#666666"),
        spaceAfter=20,
        alignment=TA_LEFT,
    )

    filter_box_style = ParagraphStyle(
        "FilterBox",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#333333"),
        leftIndent=10,
        rightIndent=10,
        spaceAfter=6,
        leading=14,
    )

    # Report title and subtitle
    story.append(Paragraph(title, report_title_style))
    story.append(Paragraph("Detailed System Report", subtitle_style))

    # Filter information in a clean box
    active_filters = {k: v for k, v in (filters or {}).items() if v}
    if active_filters:
        filter_items = []
        for key, value in active_filters.items():
            if "date" in key.lower() and value:
                filter_items.append(f"<b>{key}:</b> {value}")
            else:
                filter_items.append(f"<b>{key}:</b> {value}")

        filter_content = " &nbsp;&nbsp;|&nbsp;&nbsp; ".join(filter_items)
        story.append(Spacer(1, 10))
        story.append(
            Paragraph(
                f"<para bgcolor='#f0f0f0' borderColor='#e10600' borderWidth='1' borderPadding='8'>"
                f"<font color='#e10600'><b>Active Filters:</b></font> {filter_content}"
                f"</para>",
                filter_box_style,
            )
        )

    # Summary stats
    story.append(Spacer(1, 15))
    stats_style = ParagraphStyle(
        "Stats",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#333333"),
    )
    story.append(Paragraph(f"<b>Total Records:</b> {len(data)}", stats_style))
    story.append(Spacer(1, 20))

    # Calculate optimal column widths based on content
    col_count = len(headers)
    available_width = doc.width

    # Default even distribution with minimum width
    col_widths = [
        max(available_width / col_count, 0.8 * inch) for _ in range(col_count)
    ]

    # Adjust specific column types
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if any(
            word in header_lower for word in ["id", "time", "date", "phone", "status"]
        ):
            col_widths[i] = 0.9 * inch
        elif any(
            word in header_lower
            for word in ["name", "company", "purpose", "department", "email"]
        ):
            col_widths[i] = 1.5 * inch
        elif "position" in header_lower or "type" in header_lower:
            col_widths[i] = 1.2 * inch

    # Normalize widths to fit available space
    total_width = sum(col_widths)
    if total_width > available_width:
        scale_factor = available_width / total_width
        col_widths = [w * scale_factor for w in col_widths]

    # Create table with data
    table_data = [headers] + data
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Professional table styling
    table.setStyle(
        TableStyle(
            [
                # Header styling
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e10600")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("TOPPADDING", (0, 0), (-1, 0), 12),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                # Body styling with alternating colors
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f8f8f8")],
                ),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
                ("TOPPADDING", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                # Grid styling - subtle
                ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#e10600")),
                ("LINEABOVE", (0, 0), (-1, 0), 0, colors.transparent),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#cccccc")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
                # Left alignment for text columns, center for others
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ]
        )
    )

    story.append(table)

    # Build document with custom canvas
    logo_path = os.path.join(os.path.dirname(__file__), "static", "logo-dark.png")
    header_footer = HeaderFooterCanvas(logo_path)

    def add_header_footer(canvas, doc):
        header_footer.draw_header(canvas, doc)
        header_footer.draw_footer(canvas, doc)

    doc.build(story, onFirstPage=add_header_footer, onLaterPages=add_header_footer)

    output.seek(0)
    return output


# ------------------- PDF Export Routes -------------------
@app.route("/export/gate_logs/pdf")
@login_required
@role_required(["admin"])
def export_gate_logs_pdf():
    access_type = request.args.get("type", "")
    direction = request.args.get("direction", "")
    status = request.args.get("status", "")
    gate = request.args.get("gate", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    query = db_session.query(GateLog).order_by(GateLog.scanned_at.desc())

    if access_type:
        query = query.filter(GateLog.access_type == access_type)
    if direction:
        query = query.filter(GateLog.direction == direction)
    if status == "granted":
        query = query.filter(GateLog.access_granted.is_(True))
    elif status == "denied":
        query = query.filter(GateLog.access_granted.is_(False))
    if gate:
        query = query.filter(GateLog.gate_location == gate)
    if date_from:
        query = query.filter(
            GateLog.scanned_at >= datetime.strptime(date_from, "%Y-%m-%d")
        )
    if date_to:
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        query = query.filter(
            GateLog.scanned_at <= end_date.replace(hour=23, minute=59, second=59)
        )

    # PERFORMANCE: noload('*') prevents lazy-loading of relationships.
    # NOTE: Streaming opportunity — for very large datasets, replace .all()
    # with yield_per() + server-side cursor to avoid loading all 50k rows
    # into memory at once.
    from sqlalchemy.orm import noload as _noload

    logs = query.options(_noload("*")).limit(50000).all()

    headers = [
        "ID",
        "Type",
        "Name",
        "Direction",
        "Granted",
        "Reason",
        "Gate",
        "Scanned By",
        "Time",
    ]
    data = []
    for log in logs:
        data.append(
            [
                str(log.id),
                log.access_type,
                log.entity_name,
                log.direction,
                "Yes" if log.access_granted else "No",
                log.denial_reason or "",
                log.gate_location,
                log.scanned_by or "",
                log.scanned_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    filters = {
        "Type": access_type,
        "Direction": direction,
        "Status": status,
        "Date From": date_from,
        "Date To": date_to,
    }
    output = generate_pdf("Gate Access Logs Report", headers, data, filters)

    return send_file(
        output,
        as_attachment=True,
        download_name="gate_logs_report.pdf",
        mimetype="application/pdf",
    )


@app.route("/export/employees/pdf")
@login_required
@role_required(["admin", "manager"])
def export_employees_pdf():
    job_title = request.args.get("job_title", "")
    status = request.args.get("status", "")

    query = db_session.query(Employee)
    if job_title:
        query = query.filter(Employee.job_title == job_title)
    if status:
        query = query.filter(Employee.status == status)

    # NOTE: Streaming opportunity — for very large employee datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    employees = query.all()

    headers = [
        "ID",
        "EmpCode",
        "Initials",
        "FirstName",
        "SecondName",
        "Surname",
        "ID Number",
        "JobTitle",
        "Induction",
        "Induction Expiry",
        "Medical",
        "Medical Expiry",
        "Status",
    ]
    data = []
    for emp in employees:
        data.append(
            [
                str(emp.id),
                emp.emp_code,
                emp.initials or "",
                emp.first_name,
                emp.second_name or "",
                emp.surname,
                emp.id_number,
                emp.job_title or "",
                emp.induction or "",
                emp.induction_expiry.strftime("%Y-%m-%d")
                if emp.induction_expiry
                else "",
                emp.medical or "",
                emp.medical_expiry.strftime("%Y-%m-%d") if emp.medical_expiry else "",
                emp.status,
            ]
        )

    filters = {"JobTitle": job_title, "Status": status}
    output = generate_pdf("Employee Report", headers, data, filters)

    return send_file(
        output,
        as_attachment=True,
        download_name="employees_report.pdf",
        mimetype="application/pdf",
    )


@app.route("/export/visitors/pdf")
@login_required
@role_required(["admin", "manager"])
def export_visitors_pdf():
    status = request.args.get("status", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    query = db_session.query(Visitor)
    if status:
        query = query.filter(Visitor.status == status)
    if date_from:
        query = query.filter(
            Visitor.check_in_time >= datetime.strptime(date_from, "%Y-%m-%d")
        )
    if date_to:
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        query = query.filter(
            Visitor.check_in_time <= end_date.replace(hour=23, minute=59, second=59)
        )

    # PERFORMANCE: eager-load host relationship to avoid N+1 queries
    # when accessing visitor.host.name for each row.
    from sqlalchemy.orm import joinedload as _joinedload

    # NOTE: Streaming opportunity — for very large visitor datasets, use
    # yield_per() with a server-side cursor to stream rows instead of
    # loading all records into memory at once.
    visitors = query.options(_joinedload(Visitor.host)).all()

    headers = [
        "ID",
        "Name",
        "Company",
        "Purpose",
        "Host",
        "Check In",
        "Check Out",
        "Status",
    ]
    data = []
    for visitor in visitors:
        host_name = visitor.host.name if visitor.host else ""
        data.append(
            [
                str(visitor.id),
                visitor.name,
                visitor.company or "",
                visitor.purpose or "",
                host_name,
                visitor.check_in_time.strftime("%Y-%m-%d %H:%M")
                if visitor.check_in_time
                else "",
                visitor.check_out_time.strftime("%Y-%m-%d %H:%M")
                if visitor.check_out_time
                else "",
                visitor.status,
            ]
        )

    filters = {"Status": status, "Date From": date_from, "Date To": date_to}
    output = generate_pdf("Visitor Report", headers, data, filters)

    return send_file(
        output,
        as_attachment=True,
        download_name="visitors_report.pdf",
        mimetype="application/pdf",
    )


# ------------------- Import Routes -------------------
@app.route("/import/employees", methods=["POST"])
@login_required
@role_required(["admin", "manager"])
@limiter.limit("10 per minute")
def import_employees():
    """Import employees from Excel or CSV file."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    # Check file size (max 5MB)
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    if file_size > 5 * 1024 * 1024:
        return jsonify({"error": "File too large (max 5MB)"}), 413

    try:
        # Read file based on extension
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(file)
        else:
            return jsonify({"error": "Invalid file format. Use .csv or .xlsx"}), 400

        # Validate required columns
        required = ["emp_code", "first_name", "surname", "id_number"]
        missing = [col for col in required if col not in df.columns]
        if missing:
            return jsonify(
                {"error": f"Missing required columns: {', '.join(missing)}"}
            ), 400

        # Get existing employee codes for duplicate checking
        existing_codes = {e.emp_code for e in db_session.query(Employee.emp_code).all()}

        imported = 0
        skipped = 0
        errors = []

        # Process each row
        for idx, row in df.iterrows():
            try:
                emp_code = str(row.get("emp_code", "")).strip()
                first_name = str(row.get("first_name", "")).strip()
                surname = str(row.get("surname", "")).strip()
                id_number = str(row.get("id_number", "")).strip()

                if not emp_code or not first_name or not surname or not id_number:
                    errors.append(f"Row {idx + 2}: Missing required field")
                    continue

                # Skip duplicates
                if emp_code in existing_codes:
                    skipped += 1
                    continue

                # Parse expiry dates
                induction_expiry = None
                if "induction_expiry" in df.columns and pd.notna(
                    row.get("induction_expiry")
                ):
                    try:
                        induction_expiry = pd.to_datetime(
                            row["induction_expiry"]
                        ).to_pydatetime()
                    except Exception:
                        errors.append(f"Row {idx + 2}: Invalid induction_expiry format")

                medical_expiry = None
                if "medical_expiry" in df.columns and pd.notna(
                    row.get("medical_expiry")
                ):
                    try:
                        medical_expiry = pd.to_datetime(
                            row["medical_expiry"]
                        ).to_pydatetime()
                    except Exception:
                        errors.append(f"Row {idx + 2}: Invalid medical_expiry format")

                # Create employee
                employee = Employee(
                    emp_code=emp_code,
                    initials=str(row.get("initials", "")).strip()
                    if pd.notna(row.get("initials"))
                    else None,
                    first_name=first_name,
                    second_name=str(row.get("second_name", "")).strip()
                    if pd.notna(row.get("second_name"))
                    else None,
                    surname=surname,
                    id_number=id_number,
                    job_title=str(row.get("job_title", "")).strip()
                    if pd.notna(row.get("job_title"))
                    else None,
                    induction=str(row.get("induction", "")).strip()
                    if pd.notna(row.get("induction"))
                    else None,
                    medical=str(row.get("medical", "")).strip()
                    if pd.notna(row.get("medical"))
                    else None,
                    induction_expiry=induction_expiry,
                    medical_expiry=medical_expiry,
                    status=str(row.get("status", "Active")).strip()
                    if pd.notna(row.get("status"))
                    else "Active",
                )

                db_session.add(employee)
                db_session.flush()  # Get ID without committing

                # Generate QR code
                qr_data = f"EMP:{employee.id}:{employee.emp_code}:{datetime.now().timestamp()}"
                employee.qr_code = hashlib.sha256(qr_data.encode()).hexdigest()[:32]

                existing_codes.add(emp_code)
                imported += 1

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        db_session.commit()

        return jsonify(
            {
                "imported": imported,
                "skipped": skipped,
                "errors": errors[:10],  # Limit errors returned
                "total_rows": len(df),
            }
        )

    except Exception as e:
        db_session.rollback()
        return jsonify({"error": f"Import failed: {str(e)}"}), 500


@app.route("/import/equipment", methods=["POST"])
@login_required
@role_required(["admin", "manager"])
@limiter.limit("10 per minute")
def import_equipment():
    """Import equipment from Excel or CSV file."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    # Check file size (max 5MB)
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    if file_size > 5 * 1024 * 1024:
        return jsonify({"error": "File too large (max 5MB)"}), 413

    try:
        # Read file based on extension
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(file)
        else:
            return jsonify({"error": "Invalid file format. Use .csv or .xlsx"}), 400

        # Validate required columns - radio_id only
        required = ["radio_id"]
        missing = [col for col in required if col not in df.columns]
        if missing:
            return jsonify({"error": "Missing required column: radio_id"}), 400

        # Get existing radio IDs for duplicate checking
        existing_ids = {
            item.radio_id for item in db_session.query(Equipment.radio_id).all()
        }

        imported = 0
        skipped = 0
        errors = []

        # Process each row
        for idx, row in df.iterrows():
            try:
                radio_id = str(row.get("radio_id", "")).strip()

                if not radio_id or radio_id.lower() == "nan":
                    errors.append(f"Row {idx + 2}: Missing radio_id")
                    continue

                # Skip duplicates
                if radio_id in existing_ids:
                    skipped += 1
                    continue

                # Create equipment
                item = Equipment(
                    radio_id=radio_id,
                    status="Active",
                )

                db_session.add(item)
                db_session.flush()

                # Generate QR code
                qr_data = f"EQP:{item.id}:{item.radio_id}:{datetime.now().timestamp()}"
                item.qr_code = hashlib.sha256(qr_data.encode()).hexdigest()[:32].upper()

                existing_ids.add(radio_id)
                imported += 1

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        db_session.commit()

        return jsonify(
            {
                "imported": imported,
                "skipped": skipped,
                "errors": errors[:10],
            }
        )

    except Exception as e:
        db_session.rollback()
        return jsonify({"error": str(e)}), 500


@app.route("/import/vehicles", methods=["POST"])
@login_required
@role_required(["admin", "manager"])
@limiter.limit("10 per minute")
def import_vehicles():
    """Import vehicles from Excel or CSV file."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    # Check file size (max 5MB)
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    if file_size > 5 * 1024 * 1024:
        return jsonify({"error": "File too large (max 5MB)"}), 413

    try:
        # Read file based on extension
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(file)
        else:
            return jsonify({"error": "Invalid file format. Use .csv or .xlsx"}), 400

        # Validate required columns - fleet_id only
        required = ["fleet_id"]
        missing = [col for col in required if col not in df.columns]
        if missing:
            return jsonify({"error": "Missing required column: fleet_id"}), 400

        # Get existing fleet IDs for duplicate checking
        existing_ids = {v.fleet_id for v in db_session.query(Vehicle.fleet_id).all()}

        imported = 0
        skipped = 0
        errors = []

        # Process each row
        for idx, row in df.iterrows():
            try:
                fleet_id = str(row.get("fleet_id", "")).strip()

                if not fleet_id:
                    errors.append(f"Row {idx + 2}: Missing fleet_id")
                    continue

                # Skip duplicates
                if fleet_id in existing_ids:
                    skipped += 1
                    continue

                # Create vehicle - fleet_id only, default status to Active
                vehicle = Vehicle(
                    fleet_id=fleet_id,
                    status="Active",
                )

                db_session.add(vehicle)
                db_session.flush()

                # Generate QR code
                qr_data = (
                    f"VEH:{vehicle.id}:{vehicle.fleet_id}:{datetime.now().timestamp()}"
                )
                vehicle.qr_code = (
                    hashlib.sha256(qr_data.encode()).hexdigest()[:32].upper()
                )

                existing_ids.add(fleet_id)
                imported += 1

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        db_session.commit()

        return jsonify(
            {
                "imported": imported,
                "skipped": skipped,
                "errors": errors[:10],  # Limit errors shown
            }
        )

    except Exception as e:
        db_session.rollback()
        return jsonify({"error": str(e)}), 500


@app.route("/download/template/<entity_type>")
@login_required
@role_required(["admin", "manager"])
def download_import_template(entity_type):
    """Download CSV template for import."""
    if entity_type == "employees":
        output = io.StringIO()
        output.write(
            "emp_code,initials,first_name,second_name,surname,id_number,job_title,induction,induction_expiry,medical,medical_expiry,status\n"
        )
        output.write(
            "EMP001,JD,John,,Doe,ID123456789,Mine Supervisor,Standard Induction,2025-12-31,Class 1,2025-06-30,Active\n"
        )
        output.write(
            "EMP002,JS,Jane,,Smith,ID987654321,Equipment Operator,Advanced Induction,2026-06-30,Class 2,2026-03-31,Active\n"
        )
        output.write(
            "EMP003,BJ,Bob,,Johnson,ID456789123,Safety Officer,Safety Induction,2025-09-30,Class 1,2025-12-31,On Leave\n"
        )
        filename = "employees_template.csv"
    elif entity_type == "vehicles":
        output = io.StringIO()
        output.write("fleet_id\n")
        output.write("TRK001\n")
        output.write("EXC001\n")
        output.write("UTL001\n")
        filename = "vehicles_template.csv"
    elif entity_type == "equipment":
        output = io.StringIO()
        output.write("radio_id\n")
        output.write("RAD001\n")
        output.write("RAD002\n")
        output.write("RAD003\n")
        filename = "equipment_template.csv"
    else:
        return jsonify({"error": "Invalid template type"}), 400

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv",
    )


def get_local_ip():
    """Get the local IP address for network access."""
    try:
        # Create a socket to get the local IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "localhost"


def find_available_port(start_port=5000, end_port=5100, preferred_ports=None):
    """Find an available port in the given range.

    Args:
        start_port: Start of port range to scan
        end_port: End of port range to scan
        preferred_ports: List of preferred ports to check first

    Returns:
        tuple: (available_port, scanner_port)
    """
    # Check preferred ports first
    if preferred_ports:
        for port in preferred_ports:
            if is_port_available(port):
                scanner_port = find_scanner_port(port + 1, port + 100)
                return port, scanner_port

    # Scan range for main port
    for port in range(start_port, end_port + 1):
        if is_port_available(port):
            scanner_port = find_scanner_port(port + 1, port + 100)
            return port, scanner_port

    raise RuntimeError(f"No available ports found in range {start_port}-{end_port}")


def is_port_available(port):
    """Check if a port is available for use."""
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex(("0.0.0.0", port))
        sock.close()
        return result != 0  # Port is available if connection fails
    except Exception:
        return False


def kill_process_on_port(port):
    """Kill any process using the specified port."""
    try:
        port = int(port)
        if not (1 <= port <= 65535):
            raise ValueError(f"Invalid port: {port}")
    except (TypeError, ValueError):
        print(f"Invalid port value: {port!r}")
        return

    try:
        import signal
        import time

        print(f"Attempting to free port {port}...")

        # Method 1: Using lsof (Linux)
        try:
            result = subprocess.run(
                ["lsof", "-ti", f":{port}"], capture_output=True, text=True, timeout=5
            )
            if result.stdout.strip():
                pids = result.stdout.strip().split("\n")
                for pid in pids:
                    try:
                        os.kill(int(pid), signal.SIGTERM)
                        print(f"Killed process {pid} on port {port}")
                    except (ProcessLookupError, PermissionError):
                        # If SIGTERM fails, try SIGKILL
                        try:
                            os.kill(int(pid), signal.SIGKILL)
                            print(f"Force killed process {pid} on port {port}")
                        except Exception:
                            pass
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass

        # Method 2: Using fuser (Linux)
        try:
            result = subprocess.run(
                ["fuser", "-k", f"{port}/tcp"], capture_output=True, timeout=5
            )
            if result.returncode == 0:
                print(f"fuser killed processes on port {port}")
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass

        # Method 3: Using pkill with port pattern
        try:
            subprocess.run(["pkill", "-f", f":{port}"], capture_output=True, timeout=5)
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass

        # Method 4: Kill any python app.py processes
        try:
            subprocess.run(
                ["pkill", "-f", "python app.py"], capture_output=True, timeout=5
            )
            print("Killed python app.py processes")
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass

        # Wait a moment for processes to die
        time.sleep(3)

        # Verify port is now free
        if is_port_available(port):
            print(f"Port {port} is now available")
        else:
            print(f"Warning: Port {port} is still in use")
            # Try one more time with a different approach
            try:
                result = subprocess.run(
                    ["ss", "-tlnp"], capture_output=True, text=True, timeout=5
                )
                lines = result.stdout.split("\n")
                for line in lines:
                    if f":{port}" in line:
                        print(f"Found process on port {port}: {line}")
            except Exception:
                pass

    except Exception as e:
        print(f"Warning: Could not kill process on port {port}: {e}")


def find_scanner_port(start_port, end_port):
    """Find an available port for the scanner server."""
    for port in range(start_port, end_port + 1):
        if is_port_available(port):
            return port
    return None


def print_qr_code(url):
    """Print QR code as ASCII art in terminal."""
    qr = qrcode.QRCode(version=1, box_size=1, border=1)
    qr.add_data(url)
    qr.make(fit=True)

    CYAN = "\033[0;36m"
    NC = "\033[0m"

    print(f"\n  {CYAN}📱 Scan for mobile access:{NC}")
    matrix = qr.get_matrix()
    for row in matrix:
        line = ""
        for cell in row:
            line += "██" if cell else "  "
        print(f"  {CYAN}{line}{NC}")
    print()


# ------------------- System Monitoring -------------------


@app.before_request
def track_request():
    """Track requests for monitoring metrics."""
    request_timestamps.append(time.time())

    # Track endpoint hits
    endpoint = request.endpoint or "unknown"
    if endpoint not in metrics_history["endpoints"]:
        metrics_history["endpoints"][endpoint] = 0
    metrics_history["endpoints"][endpoint] += 1


# parse_log_line is imported from extensions


# ------------------- Teardown -------------------
@app.teardown_appcontext
def shutdown_session(exception=None):
    db_session.remove()


# ------------------- Secondary Server for Scanners (Dynamic Port) -------------------
def run_scanner_server(scanner_port, main_port):
    """Run a secondary Flask server on dynamic port for hardware scanners."""
    scanner_app = Flask(__name__)
    scanner_app.secret_key = app.secret_key

    # Add JSON error handlers
    @scanner_app.errorhandler(404)
    def json_404(error):
        return jsonify(
            {"error": "Not found", "message": "The requested resource was not found"}
        ), 404

    @scanner_app.errorhandler(405)
    def json_405(error):
        return jsonify(
            {
                "error": "Method not allowed",
                "message": "This HTTP method is not allowed for this endpoint",
            }
        ), 405

    @scanner_app.errorhandler(500)
    def json_500(error):
        return jsonify({"error": "Internal server error", "message": str(error)}), 500

    @scanner_app.route("/api/scan_qr", methods=["POST"])
    def scanner_api():
        """Forward scanner requests to the main app logic."""
        import requests

        try:
            # Forward the request to the main app on the detected port
            headers = {"X-API-Key": request.headers.get("X-API-Key", "")}
            response = requests.post(
                f"http://127.0.0.1:{main_port}/api/scan_qr",
                json=request.get_json(),
                headers=headers,
                timeout=5,
            )
            return jsonify(response.json()), response.status_code
        except requests.exceptions.ConnectionError:
            return jsonify(
                {"error": f"Main server not running on port {main_port}"}
            ), 503
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @scanner_app.route("/health", methods=["GET"])
    def health_check():
        return jsonify(
            {"status": "ok", "service": "scanner-port", "port": scanner_port}
        )

    scanner_app.run(host="0.0.0.0", port=scanner_port, threaded=True, debug=False)


def start_scanner_server(scanner_port, main_port):
    """Start the scanner server in a background thread."""
    scanner_thread = threading.Thread(
        target=run_scanner_server, args=(scanner_port, main_port), daemon=True
    )
    scanner_thread.start()


# ------------------- Scanner Discovery Service -------------------
class ScannerDiscoveryService:
    """UDP discovery service for Chainway/InfoWedge scanners."""

    def __init__(self, discovery_port=5000, response_port_range="5000-7000"):
        self.discovery_port = discovery_port
        self.port_range = response_port_range
        self.running = False
        self.thread = None
        self.local_ip = None  # Defer to start() to avoid blocking at import time
        self.auth_token = os.environ.get(
            "NETWORK_SCANNER_AUTH_TOKEN", "mine-net-scan-2024"
        )
        self.discovered_scanners = []  # Track discovered scanners

    def _get_local_ip(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except Exception:
            return "127.0.0.1"

    def run_discovery(self):
        self.running = True
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)

        try:
            sock.bind(("", self.discovery_port))
            print(
                f"✓ Scanner Discovery Service started on UDP port {self.discovery_port}"
            )

            while self.running:
                try:
                    data, addr = sock.recvfrom(1024)
                    message = data.decode("utf-8", errors="ignore").strip()

                    # Check if it's a discovery request
                    if "DISCOVER" in message.upper() or "SCANNER" in message.upper():
                        # Send server info
                        response = f"SERVER|IP:{self.local_ip}|PORTS:{self.port_range}|AUTH:{self.auth_token}|STATUS:READY"
                        sock.sendto(response.encode(), addr)
                        print(
                            f"Discovery request from {addr[0]} - responded with server info"
                        )

                        # Track discovered scanner
                        scanner_info = {
                            "ip": addr[0],
                            "port": addr[1],
                            "timestamp": datetime.now().isoformat(),
                            "message": message,
                        }

                        # Add if not already tracked
                        if not any(
                            s["ip"] == addr[0] for s in self.discovered_scanners
                        ):
                            self.discovered_scanners.append(scanner_info)

                            # Emit to dashboard
                            try:
                                socketio.emit("scanner_discovered", scanner_info)
                            except Exception:
                                pass

                except Exception as e:
                    if self.running:
                        print(f"Discovery service error: {e}")

        except Exception as e:
            print(f"Failed to start discovery service: {e}")
        finally:
            sock.close()

    def start(self):
        # Get local IP here instead of __init__ to avoid blocking socket at import time
        if self.local_ip is None:
            self.local_ip = self._get_local_ip()
        self.thread = threading.Thread(target=self.run_discovery, daemon=True)
        self.thread.start()

    def stop(self):
        self.running = False
        if self.thread:
            self.thread.join(timeout=2)

    def get_status(self):
        return {
            "running": self.running,
            "discovery_port": self.discovery_port,
            "server_ip": self.local_ip or "127.0.0.1",
            "port_range": self.port_range,
            "discovered_scanners": self.discovered_scanners[-10:],  # Last 10 scanners
        }


# ------------------- Network Scan Listener (All Ports 5000-7000) -------------------
class NetworkScanListener:
    """TCP socket server listening on ALL ports 5000-7000 for QR scans.

    Uses select() multiplexing to handle thousands of ports efficiently.
    Completely isolated from main web server (8080) and scanner server (8081-8082).
    """

    def __init__(self, start_port=5000, end_port=5050, auth_token=None):
        self.start_port = start_port
        self.end_port = end_port
        self.auth_token = auth_token or os.environ.get(
            "NETWORK_SCANNER_AUTH_TOKEN", "mine-net-scan-2024"
        )
        self.server_sockets = {}  # port -> socket mapping
        self.active_ports = []  # List of successfully bound ports
        self.running = False
        self.thread = None
        self.connections = []
        self.max_connections = 50
        self.connection_timeout = 10

    def is_local_ip(self, ip_address):
        """Check if IP is from local network."""
        local_prefixes = [
            "192.168.",
            "10.",
            "172.16.",
            "172.17.",
            "172.18.",
            "172.19.",
            "172.20.",
            "172.21.",
            "172.22.",
            "172.23.",
            "172.24.",
            "172.25.",
            "172.26.",
            "172.27.",
            "172.28.",
            "172.29.",
            "172.30.",
            "172.31.",
            "127.0.0.1",
        ]
        return any(ip_address.startswith(prefix) for prefix in local_prefixes)

    def handle_client(self, client_socket, client_address, server_port=None):
        """Handle a single client connection - supports both JSON and raw barcode data (InfoWedge/Chainway)."""
        try:
            client_socket.settimeout(self.connection_timeout)

            # Receive data
            data_parts = []
            while True:
                chunk = client_socket.recv(4096)
                if not chunk:
                    break
                data_parts.append(chunk)
                # Break on newline/CR (InfoWedge default), closing brace (JSON), or if chunk has content
                if (
                    b"\n" in chunk
                    or b"\r" in chunk
                    or chunk.strip().endswith(b"}")
                    or len(b"".join(data_parts)) > 0
                    and not chunk
                ):
                    break
                # Also break if we got a complete short payload (raw barcode)
                if len(b"".join(data_parts)) >= 2:
                    break

            if not data_parts:
                return

            raw_data = b"".join(data_parts).decode("utf-8").strip()

            # Initialize variables
            qr_code = None
            device_id = None
            direction = "IN"
            gate_location = f"Port {server_port}" if server_port else "Network Gate"
            is_raw_format = False

            # Try to parse as JSON first
            try:
                scan_data = json.loads(raw_data)
                qr_code = scan_data.get("qr_code", "").strip()
                device_id = scan_data.get("device_id", f"network:{client_address[0]}")
                direction = scan_data.get("direction", "IN").upper()
                gate_location = scan_data.get("gate_location", gate_location)
                auth_token = scan_data.get("auth_token")

                # Validate authentication for JSON format
                if auth_token and auth_token != self.auth_token:
                    response = {"success": False, "message": "Invalid authentication"}
                    client_socket.send(json.dumps(response).encode())
                    return

            except json.JSONDecodeError:
                # Not JSON - treat as raw barcode data from InfoWedge/Chainway
                is_raw_format = True
                qr_code = raw_data.strip()

                # Check for device ID prefix (e.g., "GATE1:ABC123")
                if ":" in qr_code and not qr_code.startswith("http"):
                    parts = qr_code.split(":", 1)
                    potential_device_id = parts[0].strip()
                    potential_qr = parts[1].strip()
                    # If prefix looks like a device ID (no spaces, reasonable length)
                    if (
                        len(potential_device_id) <= 20
                        and " " not in potential_device_id
                    ):
                        device_id = f"infowedge:{potential_device_id}"
                        qr_code = potential_qr
                    else:
                        device_id = f"infowedge:{client_address[0]}:{server_port}"
                else:
                    device_id = f"infowedge:{client_address[0]}:{server_port}"

                # Validate local IP for raw connections (security)
                if not self.is_local_ip(client_address[0]):
                    print(f"Rejected external raw connection from {client_address[0]}")
                    return

            # Process the scan using EXISTING function - SAME as other scans
            result = _process_qr_scan(
                qr_hash=qr_code,
                direction=direction,
                gate_location=gate_location,
                scanned_by=device_id,
                ip_address=client_address[0],
                user_agent="InfoWedge/ChainwayC66"
                if is_raw_format
                else "NetworkScanListener/1.0",
            )

            # Create response (JSON format for both)
            response = {
                "success": result["access_granted"],
                "message": result["denial_reason"]
                if not result["access_granted"]
                else "Access granted",
                "entity_type": result["entity_type"],
                "entity_name": result["entity_name"],
                "direction": direction,
                "open_gate": result["access_granted"],
            }

            # Send response (InfoWedge may ignore this, but custom apps will use it)
            try:
                client_socket.send(json.dumps(response).encode())
            except Exception:
                pass  # InfoWedge may close connection before receiving response

            # Log with appropriate prefix
            port_info = f" on port {server_port}" if server_port else ""
            source_type = "INFOWEDGE" if is_raw_format else "NETWORK"
            print(
                f"{source_type} SCAN{port_info}: {device_id} scanned '{qr_code[:40]}' - {'GRANTED' if result['access_granted'] else 'DENIED'}"
            )

            # Emit Socket.IO event for live dashboard updates
            try:
                socketio.emit(
                    "port_activity",
                    {
                        "port": server_port,
                        "device_id": device_id,
                        "qr_code": qr_code[:30] + "..."
                        if len(qr_code) > 30
                        else qr_code,
                        "access_granted": result["access_granted"],
                        "entity_name": result.get("entity_name", "Unknown"),
                        "timestamp": datetime.now().isoformat(),
                        "total_connections": len(
                            [c for c in self.connections if c.is_alive()]
                        ),
                        "source_type": "infowedge" if is_raw_format else "network",
                    },
                )
            except Exception:
                pass  # Socket.IO emission is optional

        except TimeoutError:
            try:
                client_socket.send(
                    json.dumps(
                        {"success": False, "message": "Connection timeout"}
                    ).encode()
                )
            except Exception:
                pass
        except Exception as e:
            print(f"InfoWedge/Network handler error: {e}")
            try:
                client_socket.send(
                    json.dumps({"success": False, "message": str(e)}).encode()
                )
            except Exception:
                pass
        finally:
            try:
                client_socket.close()
            except Exception:
                pass

    def run_server(self):
        """Main server loop on all sockets, falling back to select() if poll() is unavailable."""
        self.running = True

        has_poll = hasattr(select, "poll")
        if has_poll:
            poll_obj = select.poll()
            fd_to_port = {}  # Map file descriptor to port number
            for port, sock in self.server_sockets.items():
                try:
                    if sock.fileno() != -1:
                        poll_obj.register(sock, select.POLLIN)
                        fd_to_port[sock.fileno()] = port
                except Exception as e:
                    print(f"Warning: Could not register port {port}: {e}")
        else:
            print("select.poll() not available, falling back to select.select()")

        while self.running:
            try:
                if not self.server_sockets:
                    time.sleep(1)
                    continue

                ready_socks = []
                if has_poll:
                    ready = poll_obj.poll(1000)  # 1000ms = 1 second
                    for fd, event in ready:
                        if event & select.POLLIN:
                            server_port = fd_to_port.get(fd)
                            sock = self.server_sockets.get(server_port)
                            if sock:
                                ready_socks.append((sock, server_port))
                else:
                    # Filter out invalid/closed sockets before select
                    valid_sockets = {
                        s: p for p, s in self.server_sockets.items() if s.fileno() != -1
                    }
                    if not valid_sockets:
                        time.sleep(1)
                        continue
                    r, _, _ = select.select(list(valid_sockets.keys()), [], [], 1.0)
                    ready_socks = [(sock, valid_sockets[sock]) for sock in r]

                for sock, server_port in ready_socks:
                    try:
                        client_socket, client_address = sock.accept()

                        # Check connection limit
                        if len(self.connections) >= self.max_connections:
                            try:
                                client_socket.send(
                                    json.dumps(
                                        {
                                            "success": False,
                                            "message": "Server busy, too many connections",
                                        }
                                    ).encode()
                                )
                                client_socket.close()
                            except Exception:
                                pass
                            continue

                        # Handle client in new thread (same as existing scans)
                        client_thread = threading.Thread(
                            target=self.handle_client,
                            args=(client_socket, client_address, server_port),
                            daemon=True,
                        )
                        client_thread.start()
                        self.connections.append(client_thread)

                    except Exception as e:
                        if self.running:
                            print(f"Port accept error: {e}")
                        continue

                # Cleanup finished threads
                self.connections = [t for t in self.connections if t.is_alive()]

            except Exception as e:
                if self.running:
                    print(f"Network listener error: {e}")
                time.sleep(1)

    def start(self):
        """Start the network scan listener on ALL ports 5000-7000."""
        if self.running:
            return False, "Already running"

        # Try to bind to ALL ports in range
        reserved_ports = {8080, 8081, 8082}
        success_count = 0
        fail_count = 0

        for port in range(self.start_port, self.end_port + 1):
            if port in reserved_ports:
                continue

            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                sock.bind(("0.0.0.0", port))
                sock.listen(self.max_connections)
                sock.setblocking(False)
                self.server_sockets[port] = sock
                self.active_ports.append(port)
                success_count += 1
            except Exception:
                fail_count += 1
                try:
                    sock.close()
                except Exception:
                    pass

        if not self.active_ports:
            return False, "Failed to bind to any ports in range 5000-7000"

        # Start server thread
        self.thread = threading.Thread(target=self.run_server, daemon=True)
        self.thread.start()

        print(f"✓ Network Scan Listener: {success_count} ports active (5000-7000)")
        if fail_count > 0:
            print(f"  ({fail_count} ports unavailable)")

        return True, f"Listening on {success_count} ports (5000-7000)"

    def stop(self):
        """Stop the network scan listener and close all sockets."""
        self.running = False

        # Close all server sockets
        for port, sock in list(self.server_sockets.items()):
            try:
                sock.close()
            except Exception:
                pass

        self.server_sockets.clear()
        self.active_ports.clear()

        if self.thread:
            self.thread.join(timeout=2)

        return True, "Stopped"

    def get_status(self):
        """Get current status with all active ports."""
        return {
            "running": self.running,
            "port_count": len(self.active_ports),
            "ports": self.active_ports[:20] + ["..."]
            if len(self.active_ports) > 20
            else self.active_ports,
            "port_range": f"{self.start_port}-{self.end_port}",
            "active_connections": len([t for t in self.connections if t.is_alive()]),
        }

    def get_detailed_status(self):
        """Get detailed port status with connection activity for live monitoring."""

        # Sample of active ports with connection info (limit for performance)
        active_ports = []
        sample_size = min(100, len(self.active_ports))

        for port in self.active_ports[:sample_size]:
            # Count connections for this port (estimate based on recent activity)
            conn_count = 0
            for conn in self.connections:
                if conn.is_alive():
                    conn_count += 1

            active_ports.append(
                {"port": port, "status": "listening", "connections": conn_count}
            )

        # Calculate stats
        total_conns = len([c for c in self.connections if c.is_alive()])

        return {
            "running": self.running,
            "port_count": len(self.active_ports),
            "active_ports": active_ports,
            "stats": {
                "total_listening": len(self.active_ports),
                "with_connections": min(total_conns, len(active_ports)),
                "recent_scans": total_conns,
            },
        }


# Global instances — port 9100 is dedicated for C66/InfoWedge TCP raw output
network_listener = NetworkScanListener(start_port=5000, end_port=5050)
c66_tcp_listener = NetworkScanListener(start_port=9100, end_port=9100)
discovery_service = ScannerDiscoveryService()


# API Endpoints for Network Scanner Management
@app.route("/api/network_scanner/status", methods=["GET"])
@login_required
def network_scanner_status():
    """Get network scan listener status."""
    return jsonify(network_listener.get_status())


@app.route("/api/network_scanner/start", methods=["POST"])
@login_required
@role_required(["admin"])
def network_scanner_start():
    """Start the network scan listener."""
    success, message = network_listener.start()
    return jsonify({"success": success, "message": message})


@app.route("/api/network_scanner/stop", methods=["POST"])
@login_required
@role_required(["admin"])
def network_scanner_stop():
    """Stop network scan listener."""
    success, message = network_listener.stop()
    return jsonify({"success": success, "message": message})


@app.route("/scanner-config")
@login_required
@role_required(["admin"])
def scanner_configuration():
    """Configuration page for InfoWedge/Chainway scanners."""
    return render_template(
        "scanner_config.html",
        server_ip=discovery_service.local_ip,
        port_range="5000-7000",
        discovery_port=discovery_service.discovery_port,
        auth_token=discovery_service.auth_token,
    )


@app.route("/api/scanner/discovery-status")
@login_required
def scanner_discovery_status():
    """Get discovery service status and discovered scanners."""
    return jsonify(discovery_service.get_status())


@app.route("/api/scanner/test-discovery", methods=["POST"])
@login_required
@role_required(["admin"])
def test_scanner_discovery():
    """Test discovery by broadcasting a message."""
    try:
        import socket

        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)

        # Send discovery message
        message = "DISCOVER_SCANNER_SERVER"
        sock.sendto(
            message.encode(), ("255.255.255.255", discovery_service.discovery_port)
        )
        sock.close()

        return jsonify(
            {
                "success": True,
                "message": f"Discovery test sent to UDP port {discovery_service.discovery_port}",
            }
        )
    except Exception as e:
        return jsonify(
            {"success": False, "message": f"Failed to send discovery test: {str(e)}"}
        )


@app.route("/api/network_scanner/ports", methods=["GET"])
@login_required
def network_scanner_ports():
    """Get detailed port status for live monitoring dashboard."""
    return jsonify(network_listener.get_detailed_status())


def start_network_listener():
    """Auto-start network listener, C66 TCP listener, and discovery service on app startup."""
    success, message = network_listener.start()
    if success:
        print(
            f"✓ Network scan listener started on ports {network_listener.start_port}-{network_listener.end_port}"
        )
    else:
        print(f"✗ Failed to start network scan listener: {message}")

    # Dedicated port 9100 for C66 InfoWedge IP output (TCP raw)
    ok2, msg2 = c66_tcp_listener.start()
    if ok2:
        print("✓ C66 TCP listener started on port 9100")
    else:
        print(f"  C66 port 9100: {msg2}")

    try:
        discovery_service.start()
        print(
            f"✓ Scanner discovery service started on UDP port {discovery_service.discovery_port}"
        )
    except Exception as e:
        print(f"✗ Failed to start discovery service: {e}")


# Register blueprints
from routes.admin import admin_bp
from routes.ai import ai_bp
from routes.auth import auth_bp
from routes.dashboard import dashboard_bp
from routes.devices import devices_bp
from routes.employees import employees_bp
from routes.equipment import equipment_bp
from routes.fleet import fleet_bp
from routes.monitoring import monitoring_bp
from routes.scanning import scanning_bp
from routes.visitors import visitors_bp

app.register_blueprint(admin_bp)
app.register_blueprint(ai_bp)
app.register_blueprint(auth_bp)
app.register_blueprint(dashboard_bp)
app.register_blueprint(devices_bp)
app.register_blueprint(equipment_bp)
app.register_blueprint(employees_bp)
app.register_blueprint(fleet_bp)
app.register_blueprint(monitoring_bp)
app.register_blueprint(scanning_bp)
app.register_blueprint(visitors_bp)


# ------------------- Main -------------------
if __name__ == "__main__":
    import sys

    # Skip port management for now - use available ports
    print("Checking for available ports...")

    # Find available ports
    preferred_ports = [8080, 5000, 3000, 8000]
    try:
        main_port = 8080
        if not is_port_available(8080):
            main_port, scanner_port = find_available_port(
                start_port=5000, end_port=5100, preferred_ports=preferred_ports
            )
        else:
            # Try port 8081 first, then 8082, then find any available port
            if is_port_available(8081):
                scanner_port = 8081
            elif is_port_available(8082):
                scanner_port = 8082
            else:
                scanner_port = find_scanner_port(8083, 8200)
                if not scanner_port:
                    print("Warning: No available scanner port found, using 8083")
                    scanner_port = 8083
    except RuntimeError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    local_ip = get_local_ip()
    server_url = f"http://{local_ip}:{main_port}"

    # ANSI color codes
    BOLD = "\033[1m"
    GREEN = "\033[0;32m"
    YELLOW = "\033[1;33m"
    CYAN = "\033[0;36m"
    NC = "\033[0m"  # No Color

    print("=" * 55)
    print("   Arch-System - Starting")
    print("=" * 55)
    print(f"  Port: {main_port}    Scanner: {scanner_port}")
    print("-" * 55)
    print(f"  {BOLD}{GREEN}➜ Website: {CYAN}{server_url}{NC}")
    print("-" * 55)

    # Start secondary scanner server on dynamic port
    start_scanner_server(scanner_port, main_port)

    # Start network scan listener for WiFi devices (ports 5000-7000)
    start_network_listener()

    # Start multi-port scanner listeners (UDP + broadcast + optional sniffer)
    init_all_scanner_listeners()

    # Display QR code for mobile access
    print_qr_code(server_url)

    # Open browser after server starts (in a thread so server can start first)
    def open_browser():
        import time
        import webbrowser

        time.sleep(1.5)  # Wait for server to be ready
        webbrowser.open(f"http://localhost:{main_port}")

    threading.Thread(target=open_browser, daemon=True).start()

    run_kwargs = {
        "debug": False,
        "use_reloader": False,
        "host": "0.0.0.0",
        "port": 8080,
    }
    # Only pass allow_unsafe_werkzeug to the development (werkzeug) server
    if socketio.async_mode == "threading":
        run_kwargs["allow_unsafe_werkzeug"] = True

    socketio.run(app, **run_kwargs)
